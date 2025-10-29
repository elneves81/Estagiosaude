from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form, Response, Request
from fastapi.responses import JSONResponse
from fastapi.encoders import jsonable_encoder
import json
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import HTMLResponse, Response, RedirectResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from fastapi import Query
import sqlite3
import os
from datetime import datetime, date, timedelta
import re
import io
import contextlib
import csv as csv_mod
import logging
import traceback
from fastapi import Body

from . import crud, models, schemas, auth, db
from .integration import IntegrationConfig, fetch_estagios
from .db import SessionLocal, engine

# Criar tabelas
models.Base.metadata.create_all(bind=engine)

# Migração leve: garantir colunas novas em anexo2_atividades para instituicao_ensino e curso
try:
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'estagios.db')
    with sqlite3.connect(db_path) as _conn:
        _c = _conn.cursor()
        _c.execute("PRAGMA table_info(anexo2_atividades)")
        _cols = {row[1] for row in _c.fetchall()}
        if 'instituicao_ensino' not in _cols:
            _c.execute("ALTER TABLE anexo2_atividades ADD COLUMN instituicao_ensino TEXT")
        if 'curso' not in _cols:
            _c.execute("ALTER TABLE anexo2_atividades ADD COLUMN curso TEXT")
        _conn.commit()
except Exception:
    pass

app = FastAPI(title="Sistema de Estágios", version="1.0.0")

# Configurar CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5174",
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:8001",  # Adicionar a origem do servidor
        "http://localhost:8001",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

security = HTTPBearer()

# --------- RBAC helpers ---------
def require_roles(*allowed_roles: str):
    """Dependency factory to require authenticated user with allowed roles.

    Usage: current_user = Depends(require_roles('admin'))
    """
    def _dep(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(lambda: next(get_db()))):
        user = auth.get_current_user(db, credentials.credentials)
        if allowed_roles and user.tipo not in allowed_roles:
            raise HTTPException(status_code=403, detail="Acesso negado")
        return user
    return _dep

# Prometheus metrics (optional, free)
try:
    from prometheus_fastapi_instrumentator import Instrumentator
    _instrumentator = Instrumentator().instrument(app)
    # Will expose /metrics
    _instrumentator.expose(app, include_in_schema=False)
    print("Prometheus /metrics habilitado")
except Exception:
    print("Prometheus instrumentation nao habilitada (pacote ausente)")

# Migrações tolerantes (DDL simples via PRAGMA)
try:
    # Garantir coluna 'territorio' em anexo2_atividades
    import sqlite3, os
    from .models import DB_PATH
    if os.path.exists(DB_PATH):
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute("PRAGMA table_info(anexo2_atividades)")
            cols = [r[1] for r in cur.fetchall()]
            if 'territorio' not in cols:
                try:
                    cur.execute("ALTER TABLE anexo2_atividades ADD COLUMN territorio TEXT")
                    conn.commit()
                    print("✅ Coluna 'territorio' adicionada à tabela anexo2_atividades")
                except Exception as e:
                    print(f"⚠️ Falha ao adicionar coluna territorio: {e}")
except Exception as e:
    print(f"⚠️ Erro em migrações tolerantes: {e}")

# (removido) bloco antigo de montagem do frontend para evitar duplicidade

# Dependency para obter sessão do banco
def get_db():
    db_session = SessionLocal()
    try:
        yield db_session
    finally:
        db_session.close()

# Middleware para verificar colunas do banco na inicialização
@app.on_event("startup")
async def startup_event():
    """Verifica e adiciona colunas necessárias ao banco se não existirem"""
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'estagios.db')
    
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            
            # Verificar se as colunas novas existem na tabela estagios
            cursor.execute("PRAGMA table_info(estagios)")
            columns = [column[1] for column in cursor.fetchall()]
            
            # Adicionar colunas se não existirem
            new_columns = [
                ("disciplina", "TEXT"),
                ("nivel", "TEXT"), 
                ("num_estagiarios", "INTEGER"),
                ("carga_horaria", "TEXT"),
                ("salario", "TEXT"),
                ("observacoes", "TEXT"),
                ("instituicao_id", "INTEGER"),
                ("curso_id", "INTEGER"),
                ("unidade_id", "INTEGER")
            ]
            
            for col_name, col_type in new_columns:
                if col_name not in columns:
                    cursor.execute(f"ALTER TABLE estagios ADD COLUMN {col_name} {col_type}")
                    print(f"✅ Coluna {col_name} adicionada à tabela estagios")

            # Verificar coluna numero_conselho em supervisores
            cursor.execute("PRAGMA table_info(supervisores)")
            sup_columns = [column[1] for column in cursor.fetchall()]
            if "numero_conselho" not in sup_columns:
                cursor.execute("ALTER TABLE supervisores ADD COLUMN numero_conselho TEXT")
                print("✅ Coluna numero_conselho adicionada à tabela supervisores")

            # Migrar cnes/razao_social em instituicoes e unidades
            cursor.execute("PRAGMA table_info(instituicoes)")
            inst_columns = [column[1] for column in cursor.fetchall()]
            if "cnes" not in inst_columns:
                try:
                    cursor.execute("ALTER TABLE instituicoes ADD COLUMN cnes TEXT")
                    print("✅ Coluna cnes adicionada à tabela instituicoes")
                except Exception:
                    pass
            if "razao_social" not in inst_columns:
                try:
                    cursor.execute("ALTER TABLE instituicoes ADD COLUMN razao_social TEXT")
                    print("✅ Coluna razao_social adicionada à tabela instituicoes")
                except Exception:
                    pass
            cursor.execute("PRAGMA table_info(unidades)")
            uni_columns = [column[1] for column in cursor.fetchall()]
            if "cnes" not in uni_columns:
                try:
                    cursor.execute("ALTER TABLE unidades ADD COLUMN cnes TEXT")
                    print("✅ Coluna cnes adicionada à tabela unidades")
                except Exception:
                    pass
            if "razao_social" not in uni_columns:
                try:
                    cursor.execute("ALTER TABLE unidades ADD COLUMN razao_social TEXT")
                    print("✅ Coluna razao_social adicionada à tabela unidades")
                except Exception:
                    pass
            
            conn.commit()
    except Exception as e:
        print(f"⚠️ Erro na verificação do banco: {e}")

    # Migração: garantir coluna instituicao_id em usuarios para escopo por escola
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(usuarios)")
            cols = [c[1] for c in cursor.fetchall()]
            if "instituicao_id" not in cols:
                try:
                    cursor.execute("ALTER TABLE usuarios ADD COLUMN instituicao_id INTEGER")
                    conn.commit()
                    print("✅ Coluna instituicao_id adicionada à tabela usuarios")
                except Exception as e:
                    print(f"⚠️ Falha ao adicionar instituicao_id em usuarios: {e}")
    except Exception as e:
        print(f"⚠️ Erro ao verificar tabela usuarios: {e}")

    # Migrar colunas do anexo2 (tolerante a erros)
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(anexo2)")
            cols = [c[1] for c in cursor.fetchall()] if cursor.fetchall() else []
            add_cols = []
            if "status" not in cols: add_cols.append(("status","TEXT"))
            if "versao" not in cols: add_cols.append(("versao","INTEGER"))
            if "logo_url" not in cols: add_cols.append(("logo_url","TEXT"))
            for name, typ in add_cols:
                try:
                    cursor.execute(f"ALTER TABLE anexo2 ADD COLUMN {name} {typ}")
                    print(f"✅ Coluna {name} adicionada à tabela anexo2")
                except Exception:
                    pass
            conn.commit()
    except Exception as e:
        print(f"⚠️ Erro na migração de anexo2: {e}")

    # Criar tabela de versões caso não exista
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS anexo2_versions (\n                id INTEGER PRIMARY KEY AUTOINCREMENT,\n                estagio_id INTEGER NOT NULL,\n                versao INTEGER NOT NULL,\n                payload TEXT NOT NULL,\n                created_at DATETIME DEFAULT CURRENT_TIMESTAMP\n            )")
            conn.commit()
    except Exception as e:
        print(f"⚠️ Erro ao criar tabela anexo2_versions: {e}")

    # Migrar colunas label/comment em anexo2_versions
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(anexo2_versions)")
            cols = [c[1] for c in cursor.fetchall()]
            if "label" not in cols:
                try:
                    cursor.execute("ALTER TABLE anexo2_versions ADD COLUMN label TEXT")
                    print("✅ Coluna label adicionada à tabela anexo2_versions")
                except Exception:
                    pass
            if "comment" not in cols:
                try:
                    cursor.execute("ALTER TABLE anexo2_versions ADD COLUMN comment TEXT")
                    print("✅ Coluna comment adicionada à tabela anexo2_versions")
                except Exception:
                    pass
            conn.commit()
    except Exception as e:
        print(f"⚠️ Erro ao migrar tabela anexo2_versions: {e}")

    # Garantir curso "Técnico em Enfermagem" existente
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS cursos (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT NOT NULL, created_at DATETIME DEFAULT CURRENT_TIMESTAMP)")
            cursor.execute("SELECT id FROM cursos WHERE LOWER(nome) IN (?, ?, ?)", (
                'técnico em enfermagem', 'tecnico em enfermagem', 'tec enfermagem'
            ))
            row = cursor.fetchone()
            if not row:
                cursor.execute("INSERT INTO cursos (nome) VALUES (?)", ("Técnico em Enfermagem",))
                print("✅ Curso 'Técnico em Enfermagem' criado")
            conn.commit()
    except Exception as e:
        print(f"⚠️ Não foi possível garantir o curso padrão: {e}")

    # Criar tabela instituicao_cursos (relação oferta de cursos por instituição de ensino)
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS instituicao_cursos (id INTEGER PRIMARY KEY AUTOINCREMENT, instituicao_id INTEGER NOT NULL, curso_id INTEGER NOT NULL, nivel TEXT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP)")
            conn.commit()
    except Exception as e:
        print(f"⚠️ Erro ao criar tabela instituicao_cursos: {e}")

# Rotas principais
@app.get("/")
async def root():
    return {"name": "Sistema de Estágios", "status": "online"}

@app.get("/health")
async def health_check():
    return {"status": "healthy", "message": "API funcionando"}

# Padrões de saúde (para monitoração/DevOps)
@app.get("/livez", include_in_schema=False)
async def livez():
    return {"status": "ok"}

@app.get("/readyz", include_in_schema=False)
async def readyz():
    # Checa acesso básico ao banco
    try:
        with sqlite3.connect(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'estagios.db')) as conn:
            conn.execute("SELECT 1")
        return {"status": "ready"}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"not ready: {e}")

# Endpoint de debug (somente DEV)
@app.get("/debug/info")
async def debug_info(db: Session = Depends(get_db)):
    try:
        from . import models as _models
        # Reportar caminho do arquivo SQLite e listar usuários
        db_url = str(_models.DATABASE_URL) if hasattr(_models, "DATABASE_URL") else "n/a"
        users = db.query(_models.Usuario).all()
        return {
            "db_url": db_url,
            "users": [{"id": u.id, "email": u.email, "tipo": u.tipo, "ativo": u.ativo} for u in users],
        }
    except Exception as e:
        return {"error": str(e)}

# Autenticação
@app.post("/auth/login", response_model=schemas.Token)
async def login(user_credentials: schemas.UsuarioLogin, db: Session = Depends(get_db)):
    user = auth.authenticate_user(db, user_credentials.email.lower().strip(), user_credentials.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Credenciais inválidas",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token = auth.create_access_token(data={"sub": user.email, "tipo": user.tipo})
    return {"access_token": access_token, "token_type": "bearer", "user_type": user.tipo}

@app.get("/auth/me", response_model=schemas.Usuario)
async def get_current_user_info(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    user = auth.get_current_user(db, credentials.credentials)
    return user

# Usuários
@app.get("/usuarios", response_model=List[schemas.Usuario])
async def list_users(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    current_user = auth.get_current_user(db, credentials.credentials)
    if current_user.tipo != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores podem listar usuários")
    return crud.get_users(db)

@app.post("/auth/register", response_model=schemas.Usuario)
async def register_user(user: schemas.UsuarioCreate, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    current_user = auth.get_current_user(db, credentials.credentials)
    if current_user.tipo != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores podem criar usuários")
    
    db_user = crud.get_user_by_email(db, email=user.email)
    if db_user:
        raise HTTPException(status_code=400, detail="E-mail já cadastrado")
    return crud.create_user(db=db, user=user)

@app.put("/usuarios/{usuario_id}", response_model=schemas.Usuario)
async def update_usuario(
    usuario_id: int,
    payload: schemas.UsuarioUpdate,
    current_user=Depends(require_roles('admin')),
    db: Session = Depends(get_db)
):
    # Admin-only; allow editing any user
    try:
        updated = crud.update_user(db, usuario_id, payload)
    except Exception as e:
        # Checar possível violação de unicidade (email já usado)
        msg = str(e).lower()
        if 'unique' in msg or 'integrity' in msg or 'constraint' in msg:
            raise HTTPException(status_code=400, detail="E-mail já está em uso por outro usuário")
        raise
    if not updated:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return updated

@app.delete("/usuarios/{usuario_id}")
async def delete_usuario(
    usuario_id: int,
    current_user=Depends(require_roles('admin')),
    db: Session = Depends(get_db)
):
    # Prevent deleting self optionally (UI may block; API allows but you can enable this guard if desired)
    ok, err = crud.delete_user(db, usuario_id)
    if not ok:
        raise HTTPException(status_code=400, detail=err or "Falha ao remover usuário")
    return {"message": "Usuário removido"}

@app.patch("/usuarios/{usuario_id}/ativo")
async def toggle_usuario_ativo(
    usuario_id: int,
    body: dict,
    current_user=Depends(require_roles('admin')),
    db: Session = Depends(get_db)
):
    ativo = bool((body or {}).get('ativo', True))
    # Optional safeguard: evitar desativar o último admin logado
    u = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    if u.tipo == 'admin' and ativo is False:
        # Checar se haverá pelo menos um admin ativo restante
        restantes = db.query(models.Usuario).filter(models.Usuario.tipo == 'admin', models.Usuario.id != usuario_id, models.Usuario.ativo == True).count()
        if restantes == 0:
            raise HTTPException(status_code=400, detail="Não é possível desativar o último administrador ativo")
    updated, err = crud.set_user_active(db, usuario_id, ativo)
    if not updated:
        raise HTTPException(status_code=400, detail=err or "Falha ao atualizar status")
    return {"id": updated.id, "ativo": updated.ativo}

# Supervisores
@app.get("/supervisores", response_model=List[schemas.Supervisor])
async def list_supervisors(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    auth.get_current_user(db, credentials.credentials)
    return crud.get_supervisores(db)

@app.get("/supervisores/search")
async def search_supervisors(
    q: Optional[str] = None,
    limit: int = 20,
    offset: int = 0,
    sort: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    auth.get_current_user(db, credentials.credentials)
    limit = max(1, min(100, int(limit)))
    offset = max(0, int(offset))
    sort_field = None
    sort_dir = 'asc'
    if sort:
        if ':' in sort:
            sf, sd = sort.split(':', 1)
            sort_field = sf.strip()
            if sd.lower() in ('asc','desc'):
                sort_dir = sd.lower()
        else:
            sort_field = sort.strip()
    itens, total = crud.search_supervisores(db, q=q, limit=limit, offset=offset, sort_field=sort_field, sort_dir=sort_dir)
    return {"items": itens, "total": total, "limit": limit, "offset": offset, "sort": f"{sort_field}:{sort_dir}" if sort_field else None}

@app.get("/supervisores/paginado")
async def list_supervisors_paginated(
    q: Optional[str] = None,
    limit: int = 20,
    offset: int = 0,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    auth.get_current_user(db, credentials.credentials)
    limit = max(1, min(100, limit))
    offset = max(0, offset)
    items, total = crud.get_supervisores_paginated(db, q=q, limit=limit, offset=offset)
    return {"items": items, "total": total}

@app.post("/supervisores", response_model=schemas.Supervisor)
async def create_supervisor(supervisor: schemas.SupervisorCreate, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    current_user = auth.get_current_user(db, credentials.credentials)
    if current_user.tipo not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    return crud.create_supervisor(db=db, supervisor=supervisor)

@app.delete("/supervisores/{supervisor_id}")
async def delete_supervisor(supervisor_id: int, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    current_user = auth.get_current_user(db, credentials.credentials)
    if current_user.tipo not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    success = crud.delete_supervisor(db, supervisor_id)
    if not success:
        raise HTTPException(status_code=404, detail="Supervisor não encontrado")
    return {"message": "Supervisor removido com sucesso"}

@app.patch("/supervisores/{supervisor_id}", response_model=schemas.Supervisor)
async def update_supervisor(supervisor_id: int, payload: schemas.SupervisorUpdate, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    current_user = auth.get_current_user(db, credentials.credentials)
    if current_user.tipo not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    sup = crud.update_supervisor(db, supervisor_id, payload)
    if not sup:
        raise HTTPException(status_code=404, detail="Supervisor não encontrado")
    return sup

# Estágios
@app.get("/estagios", response_model=List[schemas.Estagio])
async def list_estagios(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    current_user = auth.get_current_user(db, credentials.credentials)
    return crud.get_estagios(db, user_type=current_user.tipo, user_id=current_user.id)

@app.get("/estagios/search")
async def search_estagios(
    q: Optional[str] = None,
    curso_id: Optional[int] = None,
    instituicao_id: Optional[int] = None,
    unidade_id: Optional[int] = None,
    supervisor_id: Optional[int] = None,
    limit: int = 20,
    offset: int = 0,
    sort: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(db, credentials.credentials)
    limit = max(1, min(100, int(limit)))
    offset = max(0, int(offset))
    sort_field = None
    sort_dir = 'desc'
    if sort:
        if ':' in sort:
            sf, sd = sort.split(':', 1)
            sort_field = sf.strip()
            if sd.lower() in ('asc','desc'):
                sort_dir = sd.lower()
        else:
            sort_field = sort.strip()

    cfg = IntegrationConfig()
    if cfg.enabled:
        try:
            ext = fetch_estagios(
                cfg,
                q=q,
                curso_id=curso_id,
                instituicao_id=instituicao_id,
                unidade_id=unidade_id,
                supervisor_id=supervisor_id,
                limit=limit,
                offset=offset,
                sort_field=sort_field,
                sort_dir=sort_dir,
            )
            # Já retorna no formato consumido pelo frontend
            return {"items": ext["items"], "total": ext["total"], "limit": limit, "offset": offset, "sort": f"{sort_field}:{sort_dir}" if sort_field else None, "source": "integration"}
        except Exception as e:
            # Fallback para local se integração falhar
            print(f"[integration] falha ao buscar estágios: {e}; usando base local")

    itens, total = crud.search_estagios(
        db,
        user_type=current_user.tipo,
        user_id=current_user.id,
        q=q,
        curso_id=curso_id,
        instituicao_id=instituicao_id,
        unidade_id=unidade_id,
        supervisor_id=supervisor_id,
        limit=limit,
        offset=offset,
        sort_field=sort_field,
        sort_dir=sort_dir,
    )
    itens_serializados = [schemas.Estagio.model_validate(i, from_attributes=True).model_dump(mode="json") for i in itens]
    return {"items": itens_serializados, "total": total, "limit": limit, "offset": offset, "sort": f"{sort_field}:{sort_dir}" if sort_field else None, "source": "local"}

@app.get("/estagios/paginado")
async def list_estagios_paginated(
    q: Optional[str] = None,
    instituicao_id: Optional[int] = None,
    curso_id: Optional[int] = None,
    unidade_id: Optional[int] = None,
    supervisor_id: Optional[int] = None,
    limit: int = 20,
    offset: int = 0,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(db, credentials.credentials)
    limit = max(1, min(100, limit))
    offset = max(0, offset)
    items, total = crud.get_estagios_paginated(
        db,
        user_type=current_user.tipo,
        user_id=current_user.id,
        q=q,
        instituicao_id=instituicao_id,
        curso_id=curso_id,
        unidade_id=unidade_id,
        supervisor_id=supervisor_id,
        limit=limit,
        offset=offset,
    )
    items_serializados = [schemas.Estagio.model_validate(i, from_attributes=True) for i in items]
    return {"items": items_serializados, "total": total}

@app.post("/estagios", response_model=schemas.Estagio)
async def create_estagio(estagio: schemas.EstagioCreate, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    current_user = auth.get_current_user(db, credentials.credentials)
    return crud.create_estagio(db=db, estagio=estagio)

@app.delete("/estagios/{estagio_id}")
async def delete_estagio(estagio_id: int, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    current_user = auth.get_current_user(db, credentials.credentials)
    success = crud.delete_estagio(db, estagio_id)
    if not success:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")
    return {"message": "Estágio removido com sucesso"}

@app.get("/estagios/{estagio_id}", response_model=schemas.Estagio)
async def get_estagio(estagio_id: int, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    auth.get_current_user(db, credentials.credentials)
    item = crud.get_estagio_by_id(db, estagio_id)
    if not item:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")
    return item

@app.put("/estagios/{estagio_id}", response_model=schemas.Estagio)
async def update_estagio(estagio_id: int, estagio: schemas.EstagioCreate, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    auth.get_current_user(db, credentials.credentials)
    updated = crud.update_estagio(db, estagio_id, estagio)
    if not updated:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")
    return updated

# Endpoints individuais de salário
@app.get("/estagios/{estagio_id}/salario")
async def get_estagio_salario(estagio_id: int, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    auth.get_current_user(db, credentials.credentials)
    item = crud.get_estagio_by_id(db, estagio_id)
    if not item:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")
    return {"id": item.id, "salario": item.salario}

@app.put("/estagios/{estagio_id}/salario")
async def update_estagio_salario(estagio_id: int, body: dict, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    auth.get_current_user(db, credentials.credentials)
    item = crud.get_estagio_by_id(db, estagio_id)
    if not item:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")
    salario = (body or {}).get("salario")
    try:
        item.salario = salario
        db.commit()
        return {"id": item.id, "salario": item.salario}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

# Catálogos
@app.get("/instituicoes", response_model=List[schemas.Instituicao])
async def list_instituicoes(db: Session = Depends(get_db), somente_ensino: Optional[bool] = False):
    """Lista instituições. Com somente_ensino=1 retorna apenas instituições com oferta de cursos."""
    if somente_ensino:
        try:
            data = (
                db.query(models.Instituicao)
                .join(models.InstituicaoCurso, models.InstituicaoCurso.instituicao_id == models.Instituicao.id)
                .group_by(models.Instituicao.id)
                .all()
            )
            return data
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao filtrar instituições de ensino: {e}")
    return crud.get_instituicoes(db)
@app.get("/instituicoes/search")
async def search_instituicoes(q: Optional[str] = None, limit: int = 20, offset: int = 0, db: Session = Depends(get_db)):
    try:
        items, total = crud.search_instituicoes(db, q=q, limit=limit, offset=offset)
        return {"items": items, "total": total}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/instituicoes", response_model=schemas.Instituicao)
async def create_instituicao(instituicao: schemas.InstituicaoCreate, current_user=Depends(require_roles('admin')), db: Session = Depends(get_db)):
    return crud.create_instituicao(db=db, instituicao=instituicao)

@app.delete("/instituicoes/{instituicao_id}")
async def delete_instituicao(instituicao_id: int, current_user=Depends(require_roles('admin')), db: Session = Depends(get_db)):
    success = crud.delete_instituicao(db, instituicao_id)
    if not success:
        raise HTTPException(status_code=404, detail="Instituição não encontrada")
    return {"message": "Instituição removida"}

# Instituições x Cursos (oferta)
@app.get("/instituicoes/{instituicao_id}/cursos")
async def list_cursos_da_instituicao(instituicao_id: int, db: Session = Depends(get_db)):
    try:
        with sqlite3.connect(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'estagios.db')) as conn:
            c = conn.cursor()
            c.execute("SELECT ic.id, ic.curso_id, ic.nivel, c.nome FROM instituicao_cursos ic JOIN cursos c ON c.id=ic.curso_id WHERE ic.instituicao_id=?", (instituicao_id,))
            rows = c.fetchall()
            return [{"id": r[0], "curso_id": r[1], "nivel": r[2], "curso_nome": r[3]} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/instituicoes/{instituicao_id}/cursos")
async def add_curso_a_instituicao(instituicao_id: int, payload: schemas.InstituicaoCursoCreate, current_user=Depends(require_roles('admin'))):
    if payload.instituicao_id != instituicao_id:
        raise HTTPException(400, "instituicao_id no body não bate com a rota")
    try:
        with sqlite3.connect(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'estagios.db')) as conn:
            c = conn.cursor()
            c.execute("INSERT INTO instituicao_cursos (instituicao_id, curso_id, nivel) VALUES (?, ?, ?)", (payload.instituicao_id, payload.curso_id, payload.nivel))
            conn.commit()
            return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/instituicoes/{instituicao_id}/cursos/{rel_id}")
async def remove_curso_da_instituicao(instituicao_id: int, rel_id: int, current_user=Depends(require_roles('admin'))):
    try:
        with sqlite3.connect(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'estagios.db')) as conn:
            c = conn.cursor()
            c.execute("DELETE FROM instituicao_cursos WHERE id=? AND instituicao_id=?", (rel_id, instituicao_id))
            conn.commit()
            return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/cursos", response_model=List[schemas.Curso])
async def list_cursos(db: Session = Depends(get_db)):
    return crud.get_cursos(db)
@app.get("/cursos/search")
async def search_cursos(q: Optional[str] = None, limit: int = 20, offset: int = 0, db: Session = Depends(get_db)):
    try:
        items, total = crud.search_cursos(db, q=q, limit=limit, offset=offset)
        return {"items": items, "total": total}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
@app.get("/cursos", response_model=List[schemas.Curso])
async def list_cursos(request: Request, db: Session = Depends(get_db)):
    data = crud.get_cursos(db)
    etag = crud.compute_etag([f"{c.id}:{c.nome}:{c.created_at}" for c in data])
    if request.headers.get('if-none-match') == etag:
        return Response(status_code=304)
    payload = [schemas.Curso.model_validate(c).model_dump(mode="json") for c in data]
    resp = JSONResponse(content=jsonable_encoder(payload))
    resp.headers['ETag'] = etag
    return resp

@app.post("/cursos", response_model=schemas.Curso)
async def create_curso(curso: schemas.CursoCreate, current_user=Depends(require_roles('admin')), db: Session = Depends(get_db)):
    return crud.create_curso(db=db, curso=curso)

@app.delete("/cursos/{curso_id}")
async def delete_curso(curso_id: int, current_user=Depends(require_roles('admin')), db: Session = Depends(get_db)):
    success = crud.delete_curso(db, curso_id)
    if not success:
        raise HTTPException(status_code=404, detail="Curso não encontrado")
    return {"message": "Curso removido"}

# -------------------------------------
# Utilitários: cálculo de horas e valores
# -------------------------------------

_DIA_MAP = {
    # pt-BR variantes → weekday ISO (segunda=0)
    'segunda': 0, 'segunda-feira': 0, 'seg': 0,
    'terca': 1, 'terça': 1, 'terca-feira': 1, 'terça-feira': 1, 'ter': 1,
    'quarta': 2, 'quarta-feira': 2, 'qua': 2,
    'quinta': 3, 'quinta-feira': 3, 'qui': 3,
    'sexta': 4, 'sexta-feira': 4, 'sex': 4,
}

def _parse_date_br(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    s = s.strip()
    # formatos: dd/mm/aaaa
    m = re.match(r"^(\d{1,2})\/(\d{1,2})\/(\d{4})$", s)
    if not m:
        return None
    d, mth, y = map(int, m.groups())
    try:
        return date(y, mth, d)
    except Exception:
        return None

def _parse_time_range_br(s: Optional[str]) -> Optional[tuple[tuple[int,int], tuple[int,int]]]:
    if not s:
        return None
    # normaliza separadores de forma tolerante: "às"/"AS"/"ÀS"/"a"/"-", múltiplos espaços
    try:
        import unicodedata
        x = s
        x = unicodedata.normalize('NFD', x)
        x = ''.join(ch for ch in x if unicodedata.category(ch) != 'Mn')  # remove acentos
    except Exception:
        x = s
    x = x.lower()
    x = re.sub(r"\s*-\s*", " as ", x)
    x = re.sub(r"\s*a\s*", " as ", x)
    x = re.sub(r"\s+", " ", x).strip()
    parts = re.split(r"\sas\s", x)
    if len(parts) != 2:
        return None
    def parse_hhmm(p: str) -> Optional[tuple[int,int]]:
        # aceita 08:00, 8:00, 08h00, 800
        m = re.search(r"(\d{1,2})[:h](\d{2})", p)
        hh = mm = None
        if m:
            hh = int(m.group(1)); mm = int(m.group(2))
        else:
            # tentar com apenas dígitos
            digits = ''.join(re.findall(r"\d", p))
            if len(digits) >= 4:
                hh = int(digits[:2]); mm = int(digits[2:4])
            elif len(digits) == 3:
                hh = int(digits[0]); mm = int(digits[1:])
            elif len(digits) == 2:
                hh = int(digits); mm = 0
        if hh is None or mm is None:
            return None
        hh = max(0, min(23, hh)); mm = max(0, min(59, mm))
        return (hh, mm)
    ini = parse_hhmm(parts[0]); fim = parse_hhmm(parts[1])
    if not ini or not fim:
        return None
    return ini, fim

def _weekday_from_str(s: Optional[str]) -> Optional[int]:
    if not s:
        return None
    k = s.strip().lower().replace('  ', ' ').replace('–', '-').replace('—', '-')
    k = k.replace(' feira', '')  # "Terça feira" -> "Terça"
    k = k.strip()
    return _DIA_MAP.get(k)

def _count_weekday_occurrences(dt_ini: date, dt_fim: date, weekday: int) -> int:
    if not dt_ini or not dt_fim or dt_fim < dt_ini:
        return 0
    # encontra a primeira ocorrência do weekday >= dt_ini
    days_ahead = (weekday - dt_ini.weekday()) % 7
    first = dt_ini + timedelta(days=days_ahead)
    if first > dt_fim:
        return 0
    # número de semanas entre first e dt_fim, inclusivo
    delta_days = (dt_fim - first).days
    return (delta_days // 7) + 1

def _calc_horas_individuais(a: models.Anexo2Atividade) -> float:
    # Se já veio preenchida no Excel/import, usar
    try:
        if a.carga_horaria_individual is not None:
            return float(str(a.carga_horaria_individual).replace(',', '.'))
    except Exception:
        pass
    # Caso contrário, calcular por duração * ocorrências
    dt_ini = _parse_date_br(a.data_inicio)
    dt_fim = _parse_date_br(a.data_fim)
    weekday = _weekday_from_str(a.dias_semana)
    rng = _parse_time_range_br(a.horario)
    if not (dt_ini and dt_fim and weekday is not None and rng):
        return 0.0
    (h1, m1), (h2, m2) = rng
    minutos = (h2*60 + m2) - (h1*60 + m1)
    if minutos <= 0:
        return 0.0
    horas_sessao = minutos / 60.0
    ocorr = _count_weekday_occurrences(dt_ini, dt_fim, weekday)
    return round(horas_sessao * ocorr, 2)

def _get_finance_config() -> dict:
    # leitura simples via sqlite3 direto para evitar ORM extra
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'estagios.db')
    cfg = {"valor_hora": 0.0, "arredondamento": "round", "moeda": "BRL"}
    try:
        with sqlite3.connect(db_path) as conn:
            c = conn.cursor()
            c.execute("CREATE TABLE IF NOT EXISTS finance_config (id INTEGER PRIMARY KEY CHECK (id = 1), valor_hora REAL DEFAULT 0, arredondamento TEXT DEFAULT 'round', moeda TEXT DEFAULT 'BRL')")
            conn.commit()
            c.execute("SELECT valor_hora, arredondamento, moeda FROM finance_config WHERE id=1")
            row = c.fetchone()
            if row:
                cfg["valor_hora"], cfg["arredondamento"], cfg["moeda"] = row
    except Exception:
        pass
    return cfg

def _set_finance_config(valor_hora: float, arredondamento: str = 'round', moeda: str = 'BRL'):
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'estagios.db')
    arred = arredondamento if arredondamento in ('round','ceil','floor') else 'round'
    try:
        with sqlite3.connect(db_path) as conn:
            c = conn.cursor()
            c.execute("CREATE TABLE IF NOT EXISTS finance_config (id INTEGER PRIMARY KEY CHECK (id = 1), valor_hora REAL DEFAULT 0, arredondamento TEXT DEFAULT 'round', moeda TEXT DEFAULT 'BRL')")
            c.execute("INSERT INTO finance_config (id, valor_hora, arredondamento, moeda) VALUES (1, ?, ?, ?) ON CONFLICT(id) DO UPDATE SET valor_hora=excluded.valor_hora, arredondamento=excluded.arredondamento, moeda=excluded.moeda", (float(valor_hora), arred, moeda))
            conn.commit()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar configuração financeira: {e}")

# Endpoints de configuração financeira
@app.get("/config/financeiro")
async def get_financeiro():
    return _get_finance_config()

@app.post("/config/financeiro")
async def set_financeiro(payload: dict):
    try:
        valor_hora = float(payload.get('valor_hora', 0))
        arred = str(payload.get('arredondamento', 'round'))
        moeda = str(payload.get('moeda', 'BRL'))
        _set_finance_config(valor_hora, arred, moeda)
        return _get_finance_config()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Payload inválido: {e}")

# ------------------------------
# Vagas (baseado em Anexo II)
# ------------------------------
@app.get("/vagas")
async def listar_vagas(
    q: Optional[str] = None,
    unidade: Optional[str] = None,
    supervisor: Optional[str] = None,
    dia: Optional[str] = None,
    exercicio: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    """Lista atividades do Anexo II com campo 'vagas' calculado.

    Vagas = quantidade_grupos * num_estagiarios_por_grupo
    """
    # RBAC: somente admin pode listar vagas (usuários 'escola' não devem visualizar)
    user = auth.get_current_user(db, credentials.credentials)
    if user.tipo != 'admin':
        raise HTTPException(status_code=403, detail="Apenas administradores podem visualizar vagas")

    limit = max(1, min(500, int(limit)))
    offset = max(0, int(offset))

    A = models.Anexo2Atividade
    P = models.Anexo2
    query = db.query(A).join(P, A.anexo2_id == P.id)

    # Filtros
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            (A.unidade_setor.ilike(like)) |
            (A.disciplina.ilike(like)) |
            (A.descricao.ilike(like)) |
            (A.dias_semana.ilike(like))
        )
    if unidade:
        query = query.filter(A.unidade_setor.ilike(f"%{unidade.strip()}%"))
    if supervisor:
        query = query.filter(A.supervisor_nome.ilike(f"%{supervisor.strip()}%"))
    if dia:
        query = query.filter(A.dias_semana.ilike(f"%{dia.strip()}%"))
    if exercicio:
        query = query.filter(P.exercicio.ilike(f"%{exercicio.strip()}%"))

    total = query.count()

    # Total de vagas neste filtro
    total_vagas_q = db.query(func.coalesce(func.sum((A.quantidade_grupos * A.num_estagiarios_por_grupo)), 0)).join(P, A.anexo2_id == P.id)
    if q:
        like = f"%{q.strip()}%"
        total_vagas_q = total_vagas_q.filter(
            (A.unidade_setor.ilike(like)) |
            (A.disciplina.ilike(like)) |
            (A.descricao.ilike(like)) |
            (A.dias_semana.ilike(like))
        )
    if unidade:
        total_vagas_q = total_vagas_q.filter(A.unidade_setor.ilike(f"%{unidade.strip()}%"))
    if supervisor:
        total_vagas_q = total_vagas_q.filter(A.supervisor_nome.ilike(f"%{supervisor.strip()}%"))
    if dia:
        total_vagas_q = total_vagas_q.filter(A.dias_semana.ilike(f"%{dia.strip()}%"))
    if exercicio:
        total_vagas_q = total_vagas_q.filter(P.exercicio.ilike(f"%{exercicio.strip()}%"))
    total_vagas = total_vagas_q.scalar() or 0

    atividades = (
        query
        .order_by(A.unidade_setor.asc(), A.disciplina.asc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    cfg = _get_finance_config()
    valor_hora = float(cfg.get('valor_hora') or 0)
    items = []
    for a in atividades:
        grupos = a.quantidade_grupos or 0
        por_grupo = a.num_estagiarios_por_grupo or 0
        vagas = int(grupos) * int(por_grupo)
        ch_ind = _calc_horas_individuais(a)
        valor_total = round(valor_hora * ch_ind * vagas, 2) if valor_hora and ch_ind and vagas else 0.0
        items.append({
            "id": a.id,
            "unidade_setor": a.unidade_setor,
            "disciplina": a.disciplina,
            "descricao": a.descricao,
            "nivel": a.nivel,
            "data_inicio": a.data_inicio,
            "data_fim": a.data_fim,
            "horario": a.horario,
            "dias_semana": a.dias_semana,
            "quantidade_grupos": grupos,
            "num_estagiarios_por_grupo": por_grupo,
            "carga_horaria_individual": a.carga_horaria_individual or ch_ind,
            "supervisor_nome": a.supervisor_nome,
            "supervisor_conselho": a.supervisor_conselho,
            "valor": a.valor,
            "instituicao_ensino": a.instituicao_ensino,
            "curso": a.curso,
            "valor_total": valor_total,
            "vagas": vagas,
        })

    return {
        "items": items,
        "total": total,
        "total_vagas": int(total_vagas),
        "limit": limit,
        "offset": offset,
    }

# ------------------------------
# Manutenção (admin): deduplicações e limpezas
# ------------------------------
@app.post("/maintenance/run")
async def maintenance_run(
    payload: dict = Body(...),
    current_user=Depends(require_roles('admin')),
):
    """Executa tarefas de manutenção administrativas.

    Body esperado:
      {
        "action": "unidades_sem_cnes" | "unidades_cnes_dup" | "cursos_dup" | "instituicoes_dup",
        "dry_run": true|false
      }
    """
    action = (payload or {}).get("action")
    dry_run = bool((payload or {}).get("dry_run", True))

    if action not in {"unidades_sem_cnes", "unidades_cnes_dup", "cursos_dup", "instituicoes_dup"}:
        raise HTTPException(status_code=400, detail="Ação inválida")

    # Mapear ação -> função runner
    def _runner(action_name: str):
        if action_name == "unidades_sem_cnes":
            from .remove_unidades_sem_cnes import run as r
            return r
        if action_name == "unidades_cnes_dup":
            from .remove_unidades_cnes_duplicado import run as r
            return r
        if action_name == "cursos_dup":
            from .remove_cursos_duplicados import run as r
            return r
        if action_name == "instituicoes_dup":
            from .remove_instituicoes_duplicadas import run as r
            return r
        return None

    fn = _runner(action)
    if not fn:
        raise HTTPException(status_code=400, detail="Ação não suportada")

    buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(buf):
            # Executa a rotina (logs serão capturados)
            fn(dry_run=dry_run)
    except Exception as e:
        # Ainda assim devolver log capturado e erro
        log = buf.getvalue()
        raise HTTPException(status_code=500, detail={"message": str(e), "log": log})

    log = buf.getvalue()
    return {"action": action, "dry_run": dry_run, "log": log}

# Criar Estágio a partir de uma Vaga (atividade do Anexo II)
@app.post("/vagas/{atividade_id}/criar-estagio", response_model=schemas.Estagio)
async def criar_estagio_a_partir_da_vaga(
    atividade_id: int,
    payload: dict = Body(...),
    current_user=Depends(require_roles('admin', 'escola', 'supervisor')),
    db: Session = Depends(get_db),
):
    """Cria um Estágio usando os dados de uma atividade do Anexo II.

    Corpo esperado (JSON): { nome: str, email: str, periodo?: str }
    Copia campos contextuais da vaga: disciplina, nivel, unidade, supervisor,
    e tenta resolver instituicao/curso por nome quando disponíveis.
    """
    # Buscar atividade (vaga) de origem
    atividade = db.query(models.Anexo2Atividade).filter(models.Anexo2Atividade.id == atividade_id).first()
    if not atividade:
        raise HTTPException(status_code=404, detail="Vaga (atividade) não encontrada")

    # Validar campos mínimos
    nome = (payload or {}).get("nome")
    email = (payload or {}).get("email")
    periodo = (payload or {}).get("periodo")
    if not nome or not email:
        raise HTTPException(status_code=400, detail="Campos obrigatórios faltando: 'nome' e 'email'")

    # Resolução de catálogos por nome (tolerante)
    def _find_by_nome_ci(model, nome: str):
        if not nome:
            return None
        # 1) Igualdade case-insensitive
        try:
            row = db.query(model).filter(func.lower(model.nome) == nome.lower()).first()
        except Exception:
            row = db.query(model).filter(model.nome == nome).first()
        if row:
            return row
        # 2) Contém (model.nome LIKE %nome%)
        try:
            row = db.query(model).filter(model.nome.ilike(f"%{nome}%")).order_by(func.length(model.nome).desc()).first()
            if row:
                return row
        except Exception:
            pass
        # 3) Tokenizar e tentar partes significativas (ex.: "UBS CENTRAL - UTI")
        try:
            parts = [p.strip() for p in re.split(r"[\-|–|—|\||/]", nome) if p and len(p.strip()) >= 3]
            for p in parts:
                r2 = db.query(model).filter(model.nome.ilike(f"%{p}%")).order_by(func.length(model.nome).desc()).first()
                if r2:
                    return r2
        except Exception:
            pass
        return None

    supervisor_id = None
    if atividade.supervisor_nome:
        sup = _find_by_nome_ci(models.Supervisor, atividade.supervisor_nome)
        if not sup:
            # Auto-criar supervisor se não existir (mantém padrão usado na criação de atividades)
            base = (atividade.supervisor_nome or "").lower().strip().replace(' ', '.')
            email_sup = f"{base}@supervisor.local" if base else None
            try:
                idx = 1
                while email_sup and db.query(models.Supervisor).filter(models.Supervisor.email == email_sup).first() is not None and idx < 100:
                    email_sup = f"{base}.{idx}@supervisor.local"
                    idx += 1
            except Exception:
                pass
            sup = models.Supervisor(
                nome=(atividade.supervisor_nome or '').strip(),
                email=email_sup or f"{datetime.utcnow().timestamp()}@supervisor.local",
                numero_conselho=(atividade.supervisor_conselho or None)
            )
            db.add(sup)
            try:
                db.flush()
            except Exception:
                db.rollback()
        if sup:
            supervisor_id = sup.id

    unidade_id = None
    if atividade.unidade_setor:
        unidade = _find_by_nome_ci(models.Unidade, atividade.unidade_setor)
        if unidade:
            unidade_id = unidade.id

    instituicao_id = None
    if getattr(atividade, 'instituicao_ensino', None):
        inst = _find_by_nome_ci(models.Instituicao, atividade.instituicao_ensino)
        if inst:
            instituicao_id = inst.id

    curso_id = None
    if getattr(atividade, 'curso', None):
        curso = _find_by_nome_ci(models.Curso, atividade.curso)
        if curso:
            curso_id = curso.id

    # Montar carga_horaria/salario a partir da vaga
    carga_horaria = None
    try:
        if atividade.carga_horaria_individual is not None:
            carga_horaria = str(atividade.carga_horaria_individual)
    except Exception:
        pass
    salario = None
    try:
        if atividade.valor is not None:
            salario = str(atividade.valor)
    except Exception:
        pass

    # Se período não informado, tentar compor um texto amigável
    if not periodo:
        partes = []
        if atividade.data_inicio and atividade.data_fim:
            partes.append(f"{atividade.data_inicio} a {atividade.data_fim}")
        elif atividade.data_inicio:
            partes.append(atividade.data_inicio)
        if atividade.dias_semana:
            partes.append(atividade.dias_semana)
        if atividade.horario:
            partes.append(atividade.horario)
        periodo = " | ".join([p for p in partes if p]) or None

    # Criar Estágio
    novo = schemas.EstagioCreate(
        nome=nome,
        email=email,
        periodo=periodo,
        supervisor_id=supervisor_id,
        instituicao_id=instituicao_id,
        curso_id=curso_id,
        unidade_id=unidade_id,
        disciplina=atividade.disciplina or None,
        nivel=atividade.nivel or None,
        num_estagiarios=1,
        carga_horaria=carga_horaria,
        salario=salario,
        observacoes=f"Criado a partir da vaga #{atividade.id}"
    )
    estagio = crud.create_estagio(db, novo)
    # Recarregar com relacionamentos
    return crud.get_estagio_by_id(db, estagio.id)

# Reset de vagas (admin)
@app.post("/vagas/reset")
async def reset_vagas(
    mode: str = Query("soft", description="soft: zera quantidades; hard: apaga atividades"),
    exercicio: Optional[str] = Query(None, description="Filtrar pelo exercício (ex.: 2025)"),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    """Reseta as vagas exibidas no dashboard de vagas.

    - soft: zera quantidade_grupos e num_estagiarios_por_grupo (mantém atividades)
    - hard: apaga atividades de Anexo II (cuidado!)
    """
    current_user = auth.get_current_user(db, credentials.credentials)
    if not current_user or getattr(current_user, "tipo", None) != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores podem resetar vagas")

    A = models.Anexo2Atividade
    P = models.Anexo2
    q = db.query(A).join(P, A.anexo2_id == P.id)
    if exercicio:
        q = q.filter(P.exercicio.ilike(f"%{exercicio.strip()}%"))

    affected = 0
    try:
        if mode == "hard":
            affected = q.delete(synchronize_session=False)
        else:
            affected = q.update({
                A.quantidade_grupos: 0,
                A.num_estagiarios_por_grupo: 0,
                A.carga_horaria_individual: None,
                A.valor: None,
            }, synchronize_session=False)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Falha ao resetar vagas: {e}")

    return {"mode": mode, "exercicio": exercicio, "affected": affected}

# Atualização pontual de atividade (edição inline)
@app.patch("/anexo2/atividades/{atividade_id}", response_model=schemas.Anexo2Atividade)
async def atualizar_atividade_anexo2(
    atividade_id: int,
    payload: dict = Body(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Atualiza campos editáveis de uma atividade do Anexo II.

    Campos permitidos: disciplina, descricao, nivel, unidade_setor, data_inicio,
    data_fim, horario, dias_semana, quantidade_grupos, num_estagiarios_por_grupo,
    supervisor_nome, supervisor_conselho, valor, territorio, instituicao_ensino, curso.
    """
    auth.get_current_user(db, credentials.credentials)
    atividade = db.query(models.Anexo2Atividade).filter(models.Anexo2Atividade.id == atividade_id).first()
    if not atividade:
        raise HTTPException(status_code=404, detail="Atividade não encontrada")

    allowed = {
        'disciplina', 'descricao', 'nivel', 'unidade_setor', 'data_inicio', 'data_fim',
        'horario', 'dias_semana', 'quantidade_grupos', 'num_estagiarios_por_grupo',
        'supervisor_nome', 'supervisor_conselho', 'valor', 'carga_horaria_individual', 'territorio',
        'instituicao_ensino', 'curso'
    }
    changed = False
    for k, v in (payload or {}).items():
        if k in allowed:
            # Tratar tipos/normalização específicos
            if k == 'carga_horaria_individual':
                # Permite limpar (None) para forçar recálculo automático depois
                if v is None or (isinstance(v, str) and v.strip() == ''):
                    setattr(atividade, k, None)
                else:
                    # Aceita número (float/int) ou string com vírgula/ponto
                    try:
                        # Mantém como string no banco
                        if isinstance(v, (int, float)):
                            setattr(atividade, k, str(v))
                        else:
                            # normaliza vírgula para ponto apenas para garantir numericidade
                            nv = str(v).strip()
                            _ = float(nv.replace(',', '.'))  # valida
                            setattr(atividade, k, nv)
                    except Exception:
                        # Se vier inválido, apenas armazena como string crua para não bloquear edição
                        setattr(atividade, k, str(v))
            else:
                setattr(atividade, k, v)
            changed = True
    if not changed:
        return atividade

    # Opcional: recalcular carga_horaria_individual se estiver vazia (ou deixar para cálculo dinâmico)
    try:
        # Só recalcular se o campo estiver vazio/None após a edição
        if not atividade.carga_horaria_individual:
            ch = _calc_horas_individuais(atividade)
            atividade.carga_horaria_individual = str(ch) if ch is not None else None
    except Exception:
        pass

    db.add(atividade)
    db.commit()
    db.refresh(atividade)
    return atividade


@app.post("/anexo2/atividades", response_model=schemas.Anexo2Atividade)
async def criar_atividade_anexo2(
    payload: dict = Body(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Cria uma nova atividade do Anexo II com validação de referências.
    
    Valida se a Unidade e o Supervisor existem no sistema.
    Detecta duplicatas antes de criar.
    """
    auth.get_current_user(db, credentials.credentials)
    
    # Extrai dados do payload
    unidade_setor = payload.get('unidade_setor', '').strip()
    supervisor_nome = payload.get('supervisor_nome', '').strip()
    disciplina = payload.get('disciplina', '').strip()
    horario = payload.get('horario', '').strip()
    dias_semana = payload.get('dias_semana', '').strip()
    
    # Validação 1: Verificar se Unidade existe
    if unidade_setor:
        unidade = db.query(models.Unidade).filter(models.Unidade.nome == unidade_setor).first()
        if not unidade:
            raise HTTPException(
                status_code=400, 
                detail=f"Unidade '{unidade_setor}' não encontrada no sistema. Verifique o nome ou cadastre a unidade primeiro."
            )
    
    # Validação 2: Supervisor — auto-criação se não existir
    if supervisor_nome:
        supervisor = db.query(models.Supervisor).filter(models.Supervisor.nome == supervisor_nome).first()
        if not supervisor:
            # Criar supervisor automaticamente com e-mail sintético estável
            base = supervisor_nome.lower().strip().replace(' ', '.')
            email = f"{base}@supervisor.local"
            try:
                idx = 1
                while db.query(models.Supervisor).filter(models.Supervisor.email == email).first() is not None and idx < 100:
                    email = f"{base}.{idx}@supervisor.local"
                    idx += 1
            except Exception:
                pass
            supervisor = models.Supervisor(
                nome=supervisor_nome,
                email=email,
                numero_conselho=(payload.get('supervisor_conselho') or '').strip() or None,
            )
            db.add(supervisor)
            db.flush()
        else:
            # Validar/mantém número do conselho quando fornecido
            supervisor_conselho = payload.get('supervisor_conselho', '').strip()
            # Se já houver número cadastrado, checar conflito
            if supervisor_conselho and supervisor.numero_conselho and supervisor.numero_conselho != supervisor_conselho:
                raise HTTPException(
                    status_code=400,
                    detail=f"Número do conselho '{supervisor_conselho}' não corresponde ao supervisor '{supervisor_nome}' (esperado: {supervisor.numero_conselho})"
                )
            # Se não houver número cadastrado e veio no payload, preencher
            if supervisor_conselho and not supervisor.numero_conselho:
                supervisor.numero_conselho = supervisor_conselho
                db.add(supervisor)
                try:
                    db.flush()
                except Exception:
                    db.rollback()
    
    # Validação 3: Detectar duplicatas
    if unidade_setor and disciplina and horario:
        duplicata = db.query(models.Anexo2Atividade).filter(
            models.Anexo2Atividade.unidade_setor == unidade_setor,
            models.Anexo2Atividade.disciplina == disciplina,
            models.Anexo2Atividade.horario == horario
        ).first()
        
        if duplicata:
            # Retornar aviso ao invés de erro, permitindo que o frontend decida
            if not payload.get('force_create', False):
                raise HTTPException(
                    status_code=409,
                    detail={
                        "message": "Possível vaga duplicada encontrada",
                        "existing_vaga": {
                            "id": duplicata.id,
                            "unidade": duplicata.unidade_setor,
                            "disciplina": duplicata.disciplina,
                            "horario": duplicata.horario,
                            "dias_semana": duplicata.dias_semana,
                            "supervisor": duplicata.supervisor_nome
                        }
                    }
                )
    
    # Buscar anexo2_id padrão
    # Nota: Atividades devem pertencer a um Anexo2 existente (modelo exige estagio_id)
    anexo2_id = payload.get('anexo2_id')

    if not anexo2_id:
        # Se vier um nome de estudante, criar/reutilizar um Estágio específico para ele
        estudante_nome = (payload.get('estudante_nome') or '').strip()
        if estudante_nome:
            # Reutilizar Estágio pelo nome; criar se não existir
            estagio = db.query(models.Estagio).filter(models.Estagio.nome == estudante_nome).first()
            if not estagio:
                # gerar email sintético estável a partir do nome
                base = estudante_nome.lower().replace(' ', '.')
                email = f"{base}@estudante.local"
                try:
                    idx = 1
                    while db.query(models.Estagio).filter(models.Estagio.email == email).first() is not None and idx < 100:
                        email = f"{base}.{idx}@estudante.local"
                        idx += 1
                except Exception:
                    pass
                estagio = models.Estagio(
                    nome=estudante_nome,
                    email=email,
                )
                db.add(estagio)
                db.flush()
            # Garantir Anexo II para este estágio
            anexo = db.query(models.Anexo2).filter(models.Anexo2.estagio_id == estagio.id).order_by(models.Anexo2.id.desc()).first()
            if not anexo:
                from datetime import datetime
                ano = str(datetime.utcnow().year)
                anexo = models.Anexo2(
                    estagio_id=estagio.id,
                    instituicao_ensino=payload.get('instituicao_ensino') or "",
                    curso=payload.get('curso') or "",
                    exercicio=ano,
                    status="final",
                    versao=1,
                )
                db.add(anexo)
                db.flush()
            anexo2_id = anexo.id
        else:
            # Fallback inteligente: associar ao Anexo II mais recente ou criar um 'POOL'
            ultimo_anexo = db.query(models.Anexo2).order_by(models.Anexo2.id.desc()).first()
            if not ultimo_anexo:
                # Criar (ou reutilizar) um Estágio POOL mínimo
                pool_nome = "VAGAS POOL"
                pool_email = "pool@local"
                estagio_pool = db.query(models.Estagio).filter(models.Estagio.nome == pool_nome).first()
                if not estagio_pool:
                    estagio_pool = models.Estagio(
                        nome=pool_nome,
                        email=pool_email,
                    )
                    db.add(estagio_pool)
                    db.flush()

                # Criar Anexo II padrão vinculado ao ESTÁGIO POOL
                from datetime import datetime
                ano = str(datetime.utcnow().year)
                anexo_pool = models.Anexo2(
                    estagio_id=estagio_pool.id,
                    instituicao_ensino="POOL",
                    curso="POOL",
                    exercicio=ano,
                    status="final",
                    versao=1,
                )
                db.add(anexo_pool)
                db.flush()
                anexo2_id = anexo_pool.id
            else:
                anexo2_id = ultimo_anexo.id
    
    # Criar atividade
    nova_atividade = models.Anexo2Atividade(
        anexo2_id=anexo2_id,
        disciplina=payload.get('disciplina'),
        descricao=payload.get('descricao'),
        nivel=payload.get('nivel'),
        unidade_setor=payload.get('unidade_setor'),
        data_inicio=payload.get('data_inicio'),
        data_fim=payload.get('data_fim'),
        horario=payload.get('horario'),
        dias_semana=payload.get('dias_semana'),
        quantidade_grupos=payload.get('quantidade_grupos'),
        num_estagiarios_por_grupo=payload.get('num_estagiarios_por_grupo'),
        carga_horaria_individual=payload.get('carga_horaria_individual'),
        supervisor_nome=payload.get('supervisor_nome'),
        supervisor_conselho=payload.get('supervisor_conselho'),
        valor=payload.get('valor'),
        territorio=payload.get('territorio'),
        instituicao_ensino=payload.get('instituicao_ensino'),
        curso=payload.get('curso')
    )
    
    # Calcular carga horária se não fornecida
    try:
        if not nova_atividade.carga_horaria_individual:
            ch = _calc_horas_individuais(nova_atividade)
            nova_atividade.carga_horaria_individual = str(ch) if ch is not None else None
    except Exception:
        pass
    
    db.add(nova_atividade)
    db.commit()
    db.refresh(nova_atividade)
    
    return nova_atividade

@app.delete("/anexo2/atividades/{atividade_id}")
async def deletar_atividade_anexo2(
    atividade_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Remove uma atividade do Anexo II pelo ID."""
    auth.get_current_user(db, credentials.credentials)
    atividade = db.query(models.Anexo2Atividade).filter(models.Anexo2Atividade.id == atividade_id).first()
    if not atividade:
        raise HTTPException(status_code=404, detail="Atividade não encontrada")
    db.delete(atividade)
    db.commit()
    return {"message": "Atividade removida com sucesso"}

@app.get("/anexo2/atividades/suggest-autofill")
async def sugerir_autofill(
    unidade_setor: str,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Busca última vaga da mesma unidade para auto-preencher campos.
    
    Retorna sugestões baseadas na vaga mais recente da unidade especificada.
    """
    auth.get_current_user(db, credentials.credentials)
    
    ultima_vaga = db.query(models.Anexo2Atividade).filter(
        models.Anexo2Atividade.unidade_setor == unidade_setor
    ).order_by(models.Anexo2Atividade.id.desc()).first()
    
    if not ultima_vaga:
        return {
            "found": False,
            "message": "Nenhuma vaga anterior encontrada para esta unidade"
        }
    
    return {
        "found": True,
        "suggestions": {
            "horario": ultima_vaga.horario,
            "dias_semana": ultima_vaga.dias_semana,
            "supervisor_nome": ultima_vaga.supervisor_nome,
            "supervisor_conselho": ultima_vaga.supervisor_conselho,
            "nivel": ultima_vaga.nivel,
            "quantidade_grupos": ultima_vaga.quantidade_grupos,
            "num_estagiarios_por_grupo": ultima_vaga.num_estagiarios_por_grupo,
            "valor": ultima_vaga.valor
        }
    }


# ==================== ENDPOINTS DE TEMPLATES ====================

@app.get("/vagas/templates", response_model=List[schemas.VagaTemplate])
async def listar_templates(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Lista todos os templates de vagas disponíveis."""
    user = auth.get_current_user(db, credentials.credentials)
    
    templates = db.query(models.VagaTemplate).order_by(models.VagaTemplate.nome).all()
    return templates


@app.post("/vagas/templates", response_model=schemas.VagaTemplate)
async def criar_template(
    template: schemas.VagaTemplateCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Cria um novo template de vaga."""
    user = auth.get_current_user(db, credentials.credentials)
    
    # Verificar se já existe um template com o mesmo nome
    existing = db.query(models.VagaTemplate).filter(models.VagaTemplate.nome == template.nome).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Já existe um template com o nome '{template.nome}'")
    
    novo_template = models.VagaTemplate(
        nome=template.nome,
        descricao=template.descricao,
        template_data=template.template_data,
        created_by=user.id
    )
    
    db.add(novo_template)
    db.commit()
    db.refresh(novo_template)
    
    return novo_template


@app.get("/vagas/templates/{template_id}", response_model=schemas.VagaTemplate)
async def obter_template(
    template_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Obtém os detalhes de um template específico."""
    auth.get_current_user(db, credentials.credentials)
    
    template = db.query(models.VagaTemplate).filter(models.VagaTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    return template


@app.put("/vagas/templates/{template_id}", response_model=schemas.VagaTemplate)
async def atualizar_template(
    template_id: int,
    template_update: schemas.VagaTemplateCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Atualiza um template existente."""
    auth.get_current_user(db, credentials.credentials)
    
    template = db.query(models.VagaTemplate).filter(models.VagaTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    template.nome = template_update.nome
    template.descricao = template_update.descricao
    template.template_data = template_update.template_data
    template.updated_at = datetime.utcnow()
    
    db.commit()
    db.refresh(template)
    
    return template


@app.delete("/vagas/templates/{template_id}")
async def deletar_template(
    template_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Remove um template."""
    auth.get_current_user(db, credentials.credentials)
    
    template = db.query(models.VagaTemplate).filter(models.VagaTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    db.delete(template)
    db.commit()
    
    return {"message": f"Template '{template.nome}' removido com sucesso"}


# ==================== ENDPOINT DE BULK IMPORT ====================

@app.post("/vagas/import")
async def importar_vagas_bulk(
    file: UploadFile = File(...),
    anexo2_id: Optional[int] = Form(None),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Importa múltiplas vagas de um arquivo CSV/Excel.
    
    Formato esperado (CSV com cabeçalho):
    disciplina,descricao,nivel,unidade_setor,data_inicio,data_fim,horario,dias_semana,
    quantidade_grupos,num_estagiarios_por_grupo,supervisor_nome,supervisor_conselho,valor
    """
    import io
    import csv
    import json
    
    user = auth.get_current_user(db, credentials.credentials)
    
    # Validar extensão do arquivo
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Formato de arquivo inválido. Use CSV ou Excel (.xlsx)")
    # Tratar .xls explicitamente (openpyxl não suporta .xls)
    if filename.endswith('.xls'):
        raise HTTPException(status_code=400, detail="Arquivos .xls não são suportados. Salve como .xlsx ou exporte para CSV.")
    
    results = {
        "total_linhas": 0,
        "sucesso": 0,
        "erros": [],
        "vagas_criadas": []
    }
    
    try:
        content = await file.read()
        
        # Processar CSV
        if filename.endswith('.csv'):
            # Tentar diferentes encodings e detectar delimitador ; ou ,
            text_content = None
            for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'iso-8859-1']:
                try:
                    text_content = content.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if text_content is None:
                raise HTTPException(status_code=400, detail="Não foi possível decodificar o CSV (tente UTF-8 ou Latin-1)")
            sample = text_content[:2048]
            delimiter = ';' if sample.count(';') > sample.count(',') else ','
            reader = csv.DictReader(io.StringIO(text_content), delimiter=delimiter)
            raw_headers = reader.fieldnames or []
            rows = list(reader)
        else:
            # Processar Excel
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(io.BytesIO(content))
                sheet = workbook.active
                
                # Primeira linha como cabeçalho
                headers = [cell.value for cell in sheet[1]]
                rows = []
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    row_dict = dict(zip(headers, row))
                    rows.append(row_dict)
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="Biblioteca 'openpyxl' não instalada. Instale com: pip install openpyxl"
                )
        
        # Normalizar cabeçalhos para chaves canônicas
        import unicodedata, re
        def _norm(s: str | None) -> str:
            if s is None:
                return ''
            s = str(s).strip().lower()
            s = unicodedata.normalize('NFD', s)
            s = ''.join(ch for ch in s if not unicodedata.combining(ch))
            s = s.replace('/', ' ').replace('-', ' ').replace('\\', ' ')
            s = re.sub(r'\s+', ' ', s)
            return s

        # mapa de aliases -> chave canônica
        ALIASES = {
            'disciplina': 'disciplina',
            'nome da disciplina': 'disciplina',
            'componente curricular': 'disciplina',
            'atividade': 'disciplina',

            'descricao': 'descricao',
            'descricao de atividades': 'descricao',
            'descricao de atividades (descrever no minimo cinco)': 'descricao',
            'descrição': 'descricao',
            'descrição de atividades': 'descricao',

            'nivel': 'nivel',
            'nível': 'nivel',

            'unidade': 'unidade_setor',
            'setor': 'unidade_setor',
            'unidade setor': 'unidade_setor',
            'unidade  setor': 'unidade_setor',
            'unidade/ setor': 'unidade_setor',
            'unidade/setor': 'unidade_setor',
            'unidade / setor': 'unidade_setor',
            'unidade_setor': 'unidade_setor',

            'data inicio': 'data_inicio',
            'data de inicio': 'data_inicio',
            'inicio': 'data_inicio',
            'início': 'data_inicio',

            'data fim': 'data_fim',
            'data de fim': 'data_fim',
            'fim': 'data_fim',
            'termino': 'data_fim',
            'término': 'data_fim',

            'horario': 'horario',
            'horário': 'horario',
            'hora': 'horario',

            'dias da semana': 'dias_semana',
            'dia': 'dias_semana',
            'dias': 'dias_semana',

            'quantidade de grupos': 'quantidade_grupos',
            'qtd grupos': 'quantidade_grupos',
            'grupos': 'quantidade_grupos',
            'qtde grupos': 'quantidade_grupos',

            'n de estagiarios por grupo': 'num_estagiarios_por_grupo',
            'nº de estagiarios por grupo': 'num_estagiarios_por_grupo',
            'num estagiarios por grupo': 'num_estagiarios_por_grupo',
            'estagiarios por grupo': 'num_estagiarios_por_grupo',
            'n de estagiarios': 'num_estagiarios_por_grupo',

            'carga horaria individual': 'carga_horaria_individual',
            'carga horaria': 'carga_horaria_individual',
            'ch individual': 'carga_horaria_individual',
            'ch': 'carga_horaria_individual',

            'supervisor': 'supervisor_nome',
            'nome supervisor': 'supervisor_nome',
            'supervisor responsavel': 'supervisor_nome',

            'nº conselho': 'supervisor_conselho',
            'numero conselho': 'supervisor_conselho',
            'n conselho': 'supervisor_conselho',
            'conselho': 'supervisor_conselho',

            'valor': 'valor',
            'preco': 'valor',
            'preço': 'valor',
            'custo': 'valor',
        }

        # Descobrir headers crus
        if filename.endswith('.csv'):
            headers_src = raw_headers or (rows[0].keys() if rows else [])
        else:
            headers_src = headers if 'headers' in locals() else (rows[0].keys() if rows else [])

        # Mapa: header original -> chave canonica
        header_map = {}
        for h in headers_src:
            h_norm = _norm(h)
            if h_norm in ALIASES:
                header_map[h] = ALIASES[h_norm]
            else:
                # se já estiver no formato canônico
                if h_norm in ALIASES.values():
                    header_map[h] = h_norm
        
        # Reescrever linhas com chaves canônicas
        canon_rows = []
        for r in rows:
            new_r = {}
            for k, v in (r or {}).items():
                key = header_map.get(k)
                if key:
                    new_r[key] = v
            canon_rows.append(new_r)

        rows = canon_rows
        results["total_linhas"] = len(rows)
        
        # Associar as atividades importadas a um Anexo II existente
        # Observação: o modelo Anexo2 exige estagio_id (não-nulo), então não criamos um anexo "genérico".
        if anexo2_id:
            anexo_destino = db.query(models.Anexo2).filter(models.Anexo2.id == anexo2_id).first()
            if not anexo_destino:
                raise HTTPException(status_code=404, detail=f"Anexo II com id {anexo2_id} não encontrado")
        else:
            anexo_destino = db.query(models.Anexo2).order_by(models.Anexo2.id.desc()).first()
            if not anexo_destino:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Não há Anexo II existente para associar as atividades importadas. "
                        "Crie um Anexo II primeiro e tente novamente."
                    )
                )
        
        # Processar cada linha
        for idx, row in enumerate(rows, start=1):
            try:
                # Validar campos obrigatórios
                unidade_setor = str(row.get('unidade_setor', '') or '').strip()
                disciplina = str(row.get('disciplina', '') or '').strip()
                # fallback: se disciplina vier vazia mas houver descricao, usa-la
                if not disciplina:
                    disciplina = str(row.get('descricao', '') or '').strip()
                
                if not unidade_setor or not disciplina:
                    results["erros"].append({
                        "linha": idx,
                        "erro": "Campos obrigatórios faltando: unidade_setor e disciplina são necessários"
                    })
                    continue
                
                # Validar unidade
                try:
                    unidade = db.query(models.Unidade).filter(func.lower(models.Unidade.nome) == unidade_setor.lower()).first()
                except Exception:
                    unidade = db.query(models.Unidade).filter(models.Unidade.nome == unidade_setor).first()
                if not unidade:
                    results["erros"].append({
                        "linha": idx,
                        "erro": f"Unidade '{unidade_setor}' não encontrada no sistema"
                    })
                    continue
                
                # Validar supervisor se fornecido
                supervisor_nome = str(row.get('supervisor_nome', '') or '').strip()
                if supervisor_nome:
                    try:
                        supervisor = db.query(models.Supervisor).filter(func.lower(models.Supervisor.nome) == supervisor_nome.lower()).first()
                    except Exception:
                        supervisor = db.query(models.Supervisor).filter(models.Supervisor.nome == supervisor_nome).first()
                    if not supervisor:
                        results["erros"].append({
                            "linha": idx,
                            "erro": f"Supervisor '{supervisor_nome}' não encontrado no sistema"
                        })
                        continue
                
                # Criar atividade atrelada ao Anexo II de destino
                def _to_int(val):
                    if val is None:
                        return None
                    try:
                        if isinstance(val, int):
                            return val
                        s = str(val)
                        s = s.replace('.', '').replace(',', '.')
                        # extrair numeros inteiros
                        m = re.search(r'-?\d+', s)
                        return int(m.group(0)) if m else None
                    except Exception:
                        return None
                nova_atividade = models.Anexo2Atividade(
                    anexo2_id=anexo_destino.id,
                    disciplina=disciplina,
                    descricao=str(row.get('descricao', '') or '').strip() or None,
                    nivel=str(row.get('nivel', '') or '').strip() or None,
                    unidade_setor=unidade_setor,
                    data_inicio=str(row.get('data_inicio', '') or '').strip() or None,
                    data_fim=str(row.get('data_fim', '') or '').strip() or None,
                    horario=str(row.get('horario', '') or '').strip() or None,
                    dias_semana=str(row.get('dias_semana', '') or '').strip() or None,
                    quantidade_grupos=_to_int(row.get('quantidade_grupos')),
                    num_estagiarios_por_grupo=_to_int(row.get('num_estagiarios_por_grupo')),
                    carga_horaria_individual=str(row.get('carga_horaria_individual', '') or '').strip() or None,
                    supervisor_nome=supervisor_nome or None,
                    supervisor_conselho=str(row.get('supervisor_conselho', '') or '').strip() or None,
                    valor=str(row.get('valor', '') or '').strip() or None
                )
                
                # Calcular carga horária se não fornecida
                try:
                    if not nova_atividade.carga_horaria_individual:
                        ch = _calc_horas_individuais(nova_atividade)
                        nova_atividade.carga_horaria_individual = str(ch) if ch is not None else None
                except Exception:
                    pass
                
                db.add(nova_atividade)
                db.flush()
                
                results["sucesso"] += 1
                results["vagas_criadas"].append({
                    "linha": idx,
                    "id": nova_atividade.id,
                    "unidade": unidade_setor,
                    "disciplina": disciplina
                })
                
            except Exception as e:
                results["erros"].append({
                    "linha": idx,
                    "erro": str(e)
                })
        
        # Commit se houver sucessos
        if results["sucesso"] > 0:
            db.commit()
        
        return results
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")


@app.get("/dashboard/metrics")
async def dashboard_metrics(exercicio: Optional[str] = None, top: int = 5, db: Session = Depends(get_db)):
    """Resumo consolidado para o dashboard.

    Retorna contagens principais e ranking de vagas por unidade e supervisor.
    """
    try:
        top = max(1, min(20, int(top)))
    except Exception:
        top = 5

    # Contagens simples
    total_estagios = db.query(func.count(models.Estagio.id)).scalar() or 0
    total_supervisores = db.query(func.count(models.Supervisor.id)).scalar() or 0
    total_instituicoes = db.query(func.count(models.Instituicao.id)).scalar() or 0
    total_unidades = db.query(func.count(models.Unidade.id)).scalar() or 0
    total_cursos = db.query(func.count(models.Curso.id)).scalar() or 0

    # Vagas totais e rankings
    A = models.Anexo2Atividade
    P = models.Anexo2
    base = db.query(A).join(P, A.anexo2_id == P.id)
    if exercicio:
        base = base.filter(P.exercicio.ilike(f"%{exercicio.strip()}%"))
    atividades = base.all()
    vagas_total = 0
    from collections import defaultdict
    by_unidade = defaultdict(int)
    by_supervisor = defaultdict(int)
    for a in atividades:
        grupos = int(a.quantidade_grupos or 0)
        por_grupo = int(a.num_estagiarios_por_grupo or 0)
        vagas = grupos * por_grupo
        vagas_total += vagas
        key_u = a.unidade_setor or "(sem unidade)"
        key_s = a.supervisor_nome or "(sem supervisor)"
        by_unidade[key_u] += vagas
        by_supervisor[key_s] += vagas

    def top_n(d: dict, n: int):
        items = sorted(d.items(), key=lambda kv: kv[1], reverse=True)
        return [{"chave": k, "vagas": v} for k, v in items[:n]]

    return {
        "counts": {
            "estagios": int(total_estagios),
            "supervisores": int(total_supervisores),
            "instituicoes": int(total_instituicoes),
            "unidades": int(total_unidades),
            "cursos": int(total_cursos),
        },
        "vagas_total": int(vagas_total),
        "top_unidades": top_n(by_unidade, top),
        "top_supervisores": top_n(by_supervisor, top),
    }

@app.get("/vagas/resumo")
async def resumo_vagas(
    group_by: str = "unidade",
    q: Optional[str] = None,
    unidade: Optional[str] = None,
    supervisor: Optional[str] = None,
    dia: Optional[str] = None,
    exercicio: Optional[str] = None,
    top: int = 20,
    db: Session = Depends(get_db),
):
    """Resumo de vagas agregadas por chave.

    group_by:
      - unidade: por unidade_setor
      - supervisor: por supervisor_nome
      - disciplina: por disciplina
      - dia: por dias_semana
      - unidade_dia: por unidade_setor + dias_semana
    """
    top = max(1, min(100, int(top)))
    A = models.Anexo2Atividade

    # Base query
    P = models.Anexo2
    base = db.query(A).join(P, A.anexo2_id == P.id)
    if q:
        like = f"%{q.strip()}%"
        base = base.filter(
            (A.unidade_setor.ilike(like)) |
            (A.disciplina.ilike(like)) |
            (A.descricao.ilike(like)) |
            (A.dias_semana.ilike(like))
        )
    if unidade:
        base = base.filter(A.unidade_setor.ilike(f"%{unidade.strip()}%"))
    if supervisor:
        base = base.filter(A.supervisor_nome.ilike(f"%{supervisor.strip()}%"))
    if dia:
        base = base.filter(A.dias_semana.ilike(f"%{dia.strip()}%"))
    if exercicio:
        base = base.filter(P.exercicio.ilike(f"%{exercicio.strip()}%"))

    rows = base.all()

    # Aggregation in Python (simple and SQLite-agnostic)
    from collections import defaultdict
    acc = defaultdict(lambda: {"atividades": 0, "vagas": 0})

    def key_for(a):
        if group_by == "supervisor":
            return (a.supervisor_nome or "(sem supervisor)")
        if group_by == "disciplina":
            return (a.disciplina or "(sem disciplina)")
        if group_by == "dia":
            return (a.dias_semana or "(sem dia)")
        if group_by == "unidade_dia":
            return f"{a.unidade_setor or '(sem unidade)'} | {a.dias_semana or '(sem dia)'}"
        # default: unidade
        return (a.unidade_setor or "(sem unidade)")

    for a in rows:
        k = key_for(a)
        vagas = int(a.quantidade_grupos or 0) * int(a.num_estagiarios_por_grupo or 0)
        acc[k]["atividades"] += 1
        acc[k]["vagas"] += vagas

    # Order by vagas desc
    ordered = sorted(({"chave": k, **v} for k, v in acc.items()), key=lambda x: x["vagas"], reverse=True)
    return {
        "group_by": group_by,
        "items": ordered[:top],
        "total_grupos": len(ordered),
        "total_vagas": sum(x["vagas"] for x in ordered),
    }

# Export CSV das vagas com os mesmos filtros
@app.get("/vagas/csv")
async def vagas_csv(
    q: Optional[str] = None,
    unidade: Optional[str] = None,
    supervisor: Optional[str] = None,
    dia: Optional[str] = None,
    exercicio: Optional[str] = None,
    db: Session = Depends(get_db),
):
    # Reutiliza a lógica de listar_vagas (sem paginação)
    A = models.Anexo2Atividade
    P = models.Anexo2
    query = db.query(A).join(P, A.anexo2_id == P.id)

    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            (A.unidade_setor.ilike(like)) |
            (A.disciplina.ilike(like)) |
            (A.descricao.ilike(like)) |
            (A.dias_semana.ilike(like))
        )
    if unidade:
        query = query.filter(A.unidade_setor.ilike(f"%{unidade.strip()}%"))
    if supervisor:
        query = query.filter(A.supervisor_nome.ilike(f"%{supervisor.strip()}%"))
    if dia:
        query = query.filter(A.dias_semana.ilike(f"%{dia.strip()}%"))
    if exercicio:
        query = query.filter(P.exercicio.ilike(f"%{exercicio.strip()}%"))

    atividades = query.order_by(A.unidade_setor.asc(), A.disciplina.asc()).all()

    cfg = _get_finance_config()
    valor_hora = float(cfg.get('valor_hora') or 0)

    output = io.StringIO()
    writer = csv_mod.writer(output, delimiter=';')
    writer.writerow([
        'Unidade/Setor','Disciplina','Descrição','Dia','Horário','Supervisor','Grupos','Por grupo','Vagas','CH Individual','Valor total'
    ])
    for a in atividades:
        grupos = a.quantidade_grupos or 0
        por_grupo = a.num_estagiarios_por_grupo or 0
        vagas = int(grupos) * int(por_grupo)
        ch_ind = _calc_horas_individuais(a)
        valor_total = round(valor_hora * ch_ind * vagas, 2) if valor_hora and ch_ind and vagas else 0.0
        writer.writerow([
            a.unidade_setor or '', a.disciplina or '', a.descricao or '', a.dias_semana or '', a.horario or '',
            a.supervisor_nome or '', grupos, por_grupo, vagas, ch_ind, f"{valor_total:.2f}"
        ])

    content = output.getvalue()
    return Response(content, media_type='text/csv', headers={'Content-Disposition': 'attachment; filename="vagas.csv"'})

@app.get("/unidades", response_model=List[schemas.Unidade])
async def list_unidades(request: Request, db: Session = Depends(get_db)):
    data = crud.get_unidades(db)
    etag = crud.compute_etag([f"{u.id}:{u.nome}:{u.created_at}" for u in data])
    if request.headers.get('if-none-match') == etag:
        return Response(status_code=304)
    payload = [schemas.Unidade.model_validate(u).model_dump(mode="json") for u in data]
    resp = JSONResponse(content=jsonable_encoder(payload))
    resp.headers['ETag'] = etag
    return resp
@app.get("/instituicoes", response_model=List[schemas.Instituicao])
async def list_instituicoes(request: Request, db: Session = Depends(get_db), somente_ensino: Optional[bool] = False):
    if somente_ensino:
        try:
            data = (
                db.query(models.Instituicao)
                .join(models.InstituicaoCurso, models.InstituicaoCurso.instituicao_id == models.Instituicao.id)
                .group_by(models.Instituicao.id)
                .all()
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao filtrar instituições de ensino: {e}")
    else:
        data = crud.get_instituicoes(db)
    etag = crud.compute_etag([f"{i.id}:{i.nome}:{i.created_at}" for i in data])
    if request.headers.get('if-none-match') == etag:
        return Response(status_code=304)
    payload = [schemas.Instituicao.model_validate(i).model_dump(mode="json") for i in data]
    resp = JSONResponse(content=jsonable_encoder(payload))
    resp.headers['ETag'] = etag
    return resp

@app.post("/unidades", response_model=schemas.Unidade)
async def create_unidade(unidade: schemas.UnidadeCreate, current_user=Depends(require_roles('admin')), db: Session = Depends(get_db)):
    return crud.create_unidade(db=db, unidade=unidade)

@app.delete("/unidades/{unidade_id}")
async def delete_unidade(unidade_id: int, current_user=Depends(require_roles('admin')), db: Session = Depends(get_db)):
    success = crud.delete_unidade(db, unidade_id)
    if not success:
        raise HTTPException(status_code=404, detail="Unidade não encontrada")
    return {"message": "Unidade removida"}

# Relatórios - Anexo II (HTML/PDF)
@app.get("/relatorios/anexo2/{estagio_id}")
async def relatorio_anexo2(
    estagio_id: int,
    format: Optional[str] = "html",
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    """Gera o relatório Anexo II de um estágio em HTML ou PDF.

    Query params:
    - format: 'html' (default) ou 'pdf'
    """
    # Verifica usuário
    auth.get_current_user(db, credentials.credentials)

    # Busca estágio
    estagio = crud.get_estagio_by_id(db, estagio_id)
    if not estagio:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")

    # Renderização
    try:
        from .utils_pdf import render_anexo2_html, render_anexo2
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao carregar utilitário de relatório: {e}")

    fmt = (format or "html").lower()
    if fmt == "pdf":
        try:
            anexo = crud.get_anexo2_by_estagio(db, estagio_id)
            pdf_bytes = render_anexo2(estagio, anexo2=anexo)
            return Response(
                content=pdf_bytes,
                media_type="application/pdf",
                headers={
                    # Alinhar com o nome esperado pelo frontend
                    "Content-Disposition": f"attachment; filename=plano-atividades-estagio-{estagio_id}.pdf"
                },
            )
        except ImportError as e:
            # Fallback para HTML quando WeasyPrint não está instalado
            anexo = crud.get_anexo2_by_estagio(db, estagio_id)
            html = render_anexo2_html(estagio, anexo2=anexo)
            return HTMLResponse(content=html)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao gerar PDF: {e}")
    else:
        anexo = crud.get_anexo2_by_estagio(db, estagio_id)
        html = render_anexo2_html(estagio, anexo2=anexo)
        return HTMLResponse(content=html)

# Importação - Preview
@app.post("/importar/preview")
async def preview_planilha(
    file: UploadFile = File(...),
    detectar_anexo2: bool = True,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    current_user = auth.get_current_user(db, credentials.credentials)
    
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser CSV, XLS ou XLSX")
    
    try:
        content = await file.read()
        headers = []
        rows = []
        
        # Processar CSV
        if file.filename.endswith('.csv'):
            import csv
            import io
            
            # Tentar diferentes encodings
            for encoding in ['utf-8', 'latin-1', 'iso-8859-1']:
                try:
                    text_content = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise HTTPException(status_code=400, detail="Não foi possível decodificar o arquivo CSV")
            
            # Detectar delimitador
            sample = text_content[:1024]
            delimiter = ';' if sample.count(';') > sample.count(',') else ','
            
            csvfile = io.StringIO(text_content)
            reader = csv.DictReader(csvfile, delimiter=delimiter)
            
            headers = reader.fieldnames or []
            for i, row in enumerate(reader):
                if i < 50:  # Limitar preview a 50 linhas
                    rows.append(dict(row))
                else:
                    break
        
        # Processar XLSX/XLS (detecção de linha de cabeçalho para Anexo2)
        elif file.filename.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl, io
                workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                sheet = workbook.active

                # Função para normalizar texto
                def norm(s: str | None):
                    if s is None:
                        return ''
                    import unicodedata, re
                    s = str(s).strip()
                    s = unicodedata.normalize('NFD', s)
                    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
                    return re.sub(r'\s+', ' ', s.lower())

                # Palavras-chave típicas do Anexo II
                KEYWORDS = {
                    'disciplina': ['disciplina'],
                    'descricao': ['descricao', 'descrição', 'descricao de atividades', 'descricao de atividades (descrever no minimo cinco)'],
                    'nivel': ['nivel'],
                    'unidade_setor': ['unidade', 'unidade/ setor', 'unidade setor', 'unidade/ setor'],
                    'data_inicio': ['inicio', 'início'],
                    'data_fim': ['fim'],
                    'horario': ['horario', 'horário'],
                    'dias_semana': ['dias da semana', 'dia', 'dias'],
                    'quantidade_grupos': ['quantidade de grupos'],
                    'num_estagiarios_por_grupo': ['nº de estagiarios por grupo', 'n de estagiarios por grupo', 'n de estagiarios'],
                    'carga_horaria_individual': ['carga horaria individual', 'carga horaria'],
                    'supervisor_nome': ['supervisor'],
                    'supervisor_conselho': ['nº conselho', 'numero conselho', 'n conselho'],
                    'valor': ['valor']
                }

                # Detectar linha de cabeçalho: procurar a primeira linha que contenha >=3 palavras-chave mapeáveis
                header_row_idx = 1
                detected = False
                for r in range(1, min(sheet.max_row, 30) + 1):
                    cells = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column+1)]
                    normed = [norm(c) for c in cells]
                    match_count = 0
                    for cell_val in normed:
                        for _, variants in KEYWORDS.items():
                            if any(cell_val.startswith(v) for v in variants if cell_val):
                                match_count += 1
                                break
                    if match_count >= 3:  # heurística
                        header_row_idx = r
                        detected = True
                        break

                raw_headers = [sheet.cell(row=header_row_idx, column=c).value for c in range(1, sheet.max_column+1)]
                headers = [str(h or f'Coluna_{i+1}') for i, h in enumerate(raw_headers)]

                # Construir sugestões de mapeamento
                suggestions = {}
                for i, h in enumerate(headers):
                    h_norm = norm(h)
                    for key, variants in KEYWORDS.items():
                        if any(h_norm.startswith(v) for v in variants if h_norm):
                            suggestions[key] = h  # nome real no arquivo
                            break

                # Extrair linhas seguintes até 50
                start_data_row = header_row_idx + 1
                for row_num, row in enumerate(sheet.iter_rows(min_row=start_data_row, values_only=True)):
                    if row_num >= 50:
                        break
                    row_data = {}
                    empty = True
                    for i, value in enumerate(row):
                        if i < len(headers):
                            val = '' if value is None else str(value)
                            if val.strip():
                                empty = False
                            row_data[headers[i]] = val
                    if not empty:
                        rows.append(row_data)

            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl não instalado. Execute: pip install openpyxl")
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Erro ao ler XLSX: {e}")

        extra = {}
        if 'suggestions' in locals():
            extra['suggestions'] = suggestions
            extra['header_row'] = header_row_idx if 'header_row_idx' in locals() else 1
            extra['detected'] = detected if 'detected' in locals() else False
        
        resp = {
            "headers": headers,
            "rows": rows,
            "total_rows": len(rows),
            "filename": file.filename
        }
        resp.update(extra)
        return resp
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro no preview: {str(e)}")

# Importação - Execução
@app.post("/importar/planilha")
async def importar_planilha_completa(
    file: UploadFile = File(...), 
    mapeamento: str = None,
    credentials: HTTPAuthorizationCredentials = Depends(security), 
    db: Session = Depends(get_db)
):
    current_user = auth.get_current_user(db, credentials.credentials)
    
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser CSV, XLS ou XLSX")
    
    if not mapeamento:
        raise HTTPException(status_code=400, detail="Mapeamento de colunas é obrigatório")
    
    try:
        import json
        mapeamento_dict = json.loads(mapeamento)
        
        if not mapeamento_dict.get('nome') or not mapeamento_dict.get('email'):
            raise HTTPException(status_code=400, detail="Nome e email são obrigatórios no mapeamento")
        
        content = await file.read()
        rows_data = []
        
        # Processar arquivo para obter dados
        if file.filename.endswith('.csv'):
            import csv
            import io
            
            # Tentar diferentes encodings
            for encoding in ['utf-8', 'latin-1', 'iso-8859-1']:
                try:
                    text_content = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            delimiter = ';' if text_content[:1024].count(';') > text_content[:1024].count(',') else ','
            csvfile = io.StringIO(text_content)
            reader = csv.DictReader(csvfile, delimiter=delimiter)
            
            rows_data = [dict(row) for row in reader]
        
        elif file.filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            import io
            
            workbook = openpyxl.load_workbook(io.BytesIO(content))
            sheet = workbook.active
            
            headers = [str(cell.value or f'Coluna_{i+1}') for i, cell in enumerate(sheet[1])]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = str(value) if value is not None else ''
                
                if any(row_data.values()):
                    rows_data.append(row_data)
        
        # Processar e importar dados
        imported = 0
        skipped = 0
        errors = []
        
        # Criar dicionários de lookup para evitar múltiplas consultas
        supervisores = {s.nome.lower(): s.id for s in crud.get_supervisores(db)}
        instituicoes = {i.nome.lower(): i.id for i in crud.get_instituicoes(db)}
        cursos = {c.nome.lower(): c.id for c in crud.get_cursos(db)}
        unidades = {u.nome.lower(): u.id for u in crud.get_unidades(db)}
        
        for i, row in enumerate(rows_data):
            try:
                # Mapear campos obrigatórios
                nome = row.get(mapeamento_dict['nome'], '').strip()
                email = row.get(mapeamento_dict['email'], '').strip()
                
                if not nome or not email:
                    skipped += 1
                    continue
                
                # Verificar se já existe
                existing = db.query(models.Estagio).filter(models.Estagio.email == email).first()
                if existing:
                    skipped += 1
                    continue
                
                # Mapear campos opcionais
                telefone = row.get(mapeamento_dict.get('telefone', ''), '').strip()
                periodo = row.get(mapeamento_dict.get('periodo', ''), '').strip()
                disciplina = row.get(mapeamento_dict.get('disciplina', ''), '').strip()
                nivel = row.get(mapeamento_dict.get('nivel', ''), '').strip()
                
                # Buscar IDs relacionados
                supervisor_id = None
                if mapeamento_dict.get('supervisor'):
                    supervisor_nome = row.get(mapeamento_dict['supervisor'], '').strip().lower()
                    supervisor_id = supervisores.get(supervisor_nome)
                
                instituicao_id = None
                if mapeamento_dict.get('instituicao'):
                    instituicao_nome = row.get(mapeamento_dict['instituicao'], '').strip().lower()
                    instituicao_id = instituicoes.get(instituicao_nome)
                    
                    # Se não encontrar, criar nova instituição
                    if not instituicao_id and instituicao_nome:
                        nova_inst = crud.create_instituicao(db, schemas.InstituicaoCreate(nome=instituicao_nome.title()))
                        instituicao_id = nova_inst.id
                        instituicoes[instituicao_nome] = nova_inst.id
                
                curso_id = None
                if mapeamento_dict.get('curso'):
                    curso_nome = row.get(mapeamento_dict['curso'], '').strip().lower()
                    curso_id = cursos.get(curso_nome)
                    
                    # Se não encontrar, criar novo curso
                    if not curso_id and curso_nome:
                        novo_curso = crud.create_curso(db, schemas.CursoCreate(nome=curso_nome.title()))
                        curso_id = novo_curso.id
                        cursos[curso_nome] = novo_curso.id
                
                unidade_id = None
                if mapeamento_dict.get('unidade'):
                    unidade_nome = row.get(mapeamento_dict['unidade'], '').strip().lower()
                    unidade_id = unidades.get(unidade_nome)
                    
                    # Se não encontrar, criar nova unidade
                    if not unidade_id and unidade_nome:
                        nova_unidade = crud.create_unidade(db, schemas.UnidadeCreate(nome=unidade_nome.title()))
                        unidade_id = nova_unidade.id
                        unidades[unidade_nome] = nova_unidade.id
                
                # Criar estágio
                estagio_data = schemas.EstagioCreate(
                    nome=nome,
                    email=email,
                    telefone=telefone,
                    periodo=periodo,
                    disciplina=disciplina,
                    nivel=nivel,
                    supervisor_id=supervisor_id,
                    instituicao_id=instituicao_id,
                    curso_id=curso_id,
                    unidade_id=unidade_id,
                    num_estagiarios=1,
                    observacoes=row.get(mapeamento_dict.get('observacoes', ''), '').strip()
                )
                
                crud.create_estagio(db=db, estagio=estagio_data)
                imported += 1
                
            except Exception as e:
                errors.append(f"Linha {i+2}: {str(e)}")
                skipped += 1
        
        response = {
            "message": f"Importação concluída! {imported} registros importados, {skipped} ignorados",
            "imported": imported,
            "skipped": skipped
        }
        
        if errors:
            response["errors"] = errors[:10]  # Mostrar apenas os primeiros 10 erros
        
        return response
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro na importação: {str(e)}")

# Importação legada (manter compatibilidade)
@app.post("/importar")
async def importar_planilha_legacy(file: UploadFile = File(...), credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    """Endpoint legado mantido para compatibilidade"""
    current_user = auth.get_current_user(db, credentials.credentials)
    
    if not file.filename.endswith(('.csv', '.xlsx')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser CSV ou XLSX")
    
    try:
        content = await file.read()
        
        # Importar CSV
        if file.filename.endswith('.csv'):
            import csv
            import io
            
            csvfile = io.StringIO(content.decode('utf-8'))
            reader = csv.DictReader(csvfile)
            
            imported = 0
            for row in reader:
                if all(k in row for k in ['nome', 'email']):
                    estagio_data = schemas.EstagioCreate(
                        nome=row['nome'],
                        email=row['email'],
                        telefone=row.get('telefone', ''),
                        periodo=row.get('periodo', ''),
                        supervisor_id=int(row['supervisor_id']) if row.get('supervisor_id') else None
                    )
                    crud.create_estagio(db=db, estagio=estagio_data)
                    imported += 1
        
        # Importar XLSX
        elif file.filename.endswith('.xlsx'):
            try:
                import openpyxl
                import io
                
                workbook = openpyxl.load_workbook(io.BytesIO(content))
                sheet = workbook.active
                
                headers = [cell.value for cell in sheet[1]]
                imported = 0
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:  # nome e email obrigatórios
                        row_data = dict(zip(headers, row))
                        estagio_data = schemas.EstagioCreate(
                            nome=row_data.get('nome'),
                            email=row_data.get('email'),
                            telefone=row_data.get('telefone', ''),
                            periodo=row_data.get('periodo', ''),
                            supervisor_id=int(row_data['supervisor_id']) if row_data.get('supervisor_id') else None
                        )
                        crud.create_estagio(db=db, estagio=estagio_data)
                        imported += 1
                        
            except ImportError:
                raise HTTPException(
                    status_code=500, 
                    detail="openpyxl não instalado. Execute: pip install openpyxl"
                )
        
        return {"message": f"{imported} registros importados com sucesso"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro na importação: {str(e)}")

# ------------------------------
# Importador Inteligente (novo)
# ------------------------------

@app.post("/importar/preview-inteligente")
async def preview_planilha_inteligente(
    file: UploadFile = File(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Preview inteligente que detecta automaticamente a estrutura da planilha"""
    current_user = auth.get_current_user(db, credentials.credentials)
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser XLSX ou XLS")
    
    try:
        # Salvar arquivo temporariamente
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_file_path = tmp_file.name
        
        try:
            # Usar o importador inteligente
            from .importador_inteligente import ImportadorInteligente
            
            importador = ImportadorInteligente()
            resultado = importador.analisar_planilha(tmp_file_path)
            
            # Adicionar informações extras
            resultado['arquivo_original'] = file.filename
            resultado['timestamp'] = datetime.now().isoformat()
            
            return resultado
            
        finally:
            # Limpar arquivo temporário
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)
                
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro no preview inteligente: {str(e)}")

@app.post("/importar/executar-inteligente")
async def executar_importacao_inteligente(
    request: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Executa a importação baseada no preview inteligente"""
    current_user = auth.get_current_user(db, credentials.credentials)
    
    try:
        estagio_id = request.get("estagio_id")
        atividades_selecionadas = request.get("atividades", [])
        criar_novo_estagio = request.get("criar_novo_estagio", False)

        # Validação básica de entrada
        if not isinstance(atividades_selecionadas, list):
            raise HTTPException(status_code=400, detail="Campo 'atividades' deve ser uma lista")
        
        if not atividades_selecionadas:
            raise HTTPException(status_code=400, detail="Nenhuma atividade selecionada")
        
        # Se deve criar novo estágio
        if criar_novo_estagio:
            primeiro_atividade = atividades_selecionadas[0]
            
            # Buscar ou criar instituição
            instituicao_nome = primeiro_atividade.get('instituicao', 'Instituição Importada')
            instituicao = db.query(models.Instituicao).filter(
                models.Instituicao.nome.ilike(f"%{instituicao_nome}%")
            ).first()
            
            if not instituicao:
                instituicao = crud.create_instituicao(db, schemas.InstituicaoCreate(nome=instituicao_nome))
            
            # Buscar ou criar curso
            curso_nome = primeiro_atividade.get('curso', 'Curso Importado')
            curso = db.query(models.Curso).filter(
                models.Curso.nome.ilike(f"%{curso_nome}%")
            ).first()
            
            if not curso:
                curso = crud.create_curso(db, schemas.CursoCreate(nome=curso_nome))
            
            # Criar estágio
            # Gerar email seguro e único o suficiente
            base_email = f"estagio.{instituicao_nome.lower().strip().replace(' ', '.')}@exemplo.com"
            email = base_email
            try:
                idx = 1
                while db.query(models.Estagio).filter(models.Estagio.email == email).first() is not None and idx < 100:
                    local, domain = base_email.split('@', 1)
                    email = f"{local}.{idx}@{domain}"
                    idx += 1
            except Exception:
                email = base_email
            estagio_data = schemas.EstagioCreate(
                nome=f"Estágio {instituicao_nome}",
                email=email,
                instituicao_id=instituicao.id,
                curso_id=curso.id,
                periodo=primeiro_atividade.get('periodo', ''),
                observacoes="Criado via importação inteligente"
            )
            
            estagio = crud.create_estagio(db, estagio_data)
            estagio_id = estagio.id
        
        elif not estagio_id:
            raise HTTPException(status_code=400, detail="estagio_id é obrigatório quando não criar novo estágio")
        
        # Verificar se o estágio existe
        estagio = None
        try:
            estagio = crud.get_estagio(db, estagio_id)
        except Exception:
            try:
                estagio = crud.get_estagio_by_id(db, estagio_id)
            except Exception:
                estagio = None
        if not estagio:
            raise HTTPException(status_code=404, detail="Estágio não encontrado")
        
        # Converter atividades para o formato do Anexo II
        atividades_converted = []
        conversao_erros = []
        for idx, ativ in enumerate(atividades_selecionadas, start=1):
            try:
                atividade = schemas.Anexo2AtividadeCreate(
                    disciplina=ativ.get("disciplina", ""),
                    descricao=ativ.get("descricao") or f"Importado: {ativ.get('disciplina', '')} - {ativ.get('supervisor', '')}",
                    nivel=ativ.get("nivel", "G"),
                    unidade_setor=ativ.get("unidade_setor") or ativ.get("instituicao", ""),
                    data_inicio=ativ.get("data_inicio", ""),
                    data_fim=ativ.get("data_fim", ""),
                    horario=ativ.get("horario", ""),
                    dias_semana=ativ.get("dias_semana", ativ.get("periodo", "")),
                    quantidade_grupos=int(ativ.get("grupos", 1) or 1),
                    num_estagiarios_por_grupo=int(ativ.get("num_estagiarios", 1) or 1),
                    carga_horaria_individual=str(ativ.get("horas_individuais", 4)),
                    supervisor_nome=ativ.get("supervisor", ""),
                    supervisor_conselho=str(ativ.get("supervisor_conselho", "")),
                    valor=str(ativ.get("valor", ativ.get("horas_calculadas", 0)))
                )
                atividades_converted.append(atividade)
            except Exception as conv_e:
                conversao_erros.append(f"Atividade #{idx}: {conv_e}")
                logging.warning(f"Falha ao converter atividade #{idx}: {conv_e}")

        if not atividades_converted:
            raise HTTPException(status_code=400, detail={
                "message": "Falha ao converter atividades",
                "errors": conversao_erros[:10]
            })
        
        # Criar o Anexo II
        anexo2_data = schemas.Anexo2Create(
            estagio_id=estagio_id,
            instituicao_ensino=estagio.instituicao.nome if estagio.instituicao else "",
            curso=estagio.curso.nome if estagio.curso else "",
            exercicio="2024",
            status="final",
            versao=1,
            atividades=atividades_converted
        )
        
        # Verificar se já existe anexo para este estágio
        existing = crud.get_anexo2_by_estagio(db, estagio_id)
        if existing:
            # Atualizar existente
            updated = crud.update_anexo2(db, existing.id, anexo2_data)
            try:
                crud.save_anexo2_version(db, estagio_id, updated, 
                                       label="Importação Inteligente", 
                                       comment="Atualizado via importador inteligente")
            except Exception as e:
                print(f"⚠️ Falha ao salvar versão: {e}")
            resultado = updated
        else:
            # Criar novo
            created = crud.create_anexo2(db, anexo2_data)
            try:
                crud.save_anexo2_version(db, estagio_id, created, 
                                       label="Importação Inteligente", 
                                       comment="Criado via importador inteligente")
            except Exception as e:
                print(f"⚠️ Falha ao salvar versão: {e}")
            resultado = created
        
        resp = {
            "message": f"Importação concluída! {len(atividades_selecionadas)} atividades importadas",
            "estagio_id": estagio_id,
            "anexo2_id": resultado.id,
            "atividades_importadas": len(atividades_selecionadas),
            "total_horas": sum(ativ.get("horas_calculadas", 0) for ativ in atividades_selecionadas)
        }
        if conversao_erros:
            resp["warnings"] = conversao_erros[:10]
        return resp
        
    except Exception as e:
        logging.error("Erro na importação inteligente: %s", e)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro na importação: {str(e)}")

@app.post("/importar/multiplas-planilhas")
async def processar_multiplas_planilhas(
    files: List[UploadFile] = File(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Processa múltiplas planilhas de uma vez e mostra preview consolidado"""
    current_user = auth.get_current_user(db, credentials.credentials)
    
    if len(files) > 10:
        raise HTTPException(status_code=400, detail="Máximo de 10 arquivos por vez")
    
    try:
        import tempfile
        import os
        from .importador_inteligente import ImportadorInteligente
        
        caminhos_temporarios = []
        
        # Salvar todos os arquivos temporariamente
        for file in files:
            if not file.filename.endswith(('.xlsx', '.xls')):
                continue
                
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                content = await file.read()
                tmp_file.write(content)
                caminhos_temporarios.append(tmp_file.name)
        
        try:
            # Processar todas as planilhas
            importador = ImportadorInteligente()
            resultado = importador.processar_multiplas_planilhas(caminhos_temporarios)
            
            # Adicionar nomes originais dos arquivos
            for i, planilha in enumerate(resultado['planilhas']):
                if i < len(files):
                    planilha['arquivo_original'] = files[i].filename
            
            resultado['timestamp'] = datetime.now().isoformat()
            
            return resultado
            
        finally:
            # Limpar arquivos temporários
            for caminho in caminhos_temporarios:
                if os.path.exists(caminho):
                    os.unlink(caminho)
                    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar planilhas: {str(e)}")

# (removido) endpoint duplicado de relatórios anexo2

# --------------------
# Frontend estático (build do Vite)
# --------------------
from pathlib import Path

WEB_DIST = Path(__file__).resolve().parents[2] / "web" / "dist"
if WEB_DIST.exists():
    # 1) Servir assets estáticos (JS/CSS) sob /app/assets
    app.mount("/app/assets", StaticFiles(directory=str(WEB_DIST / "assets")), name="assets")

    # 2) Servir favicon explicitamente
    @app.get("/app/favicon.svg", include_in_schema=False)
    async def app_favicon():
        return FileResponse(WEB_DIST / "favicon.svg")

    # 3) SPA fallback: qualquer outra rota em /app* devolve index.html
    @app.get("/app", include_in_schema=False)
    @app.get("/app/", include_in_schema=False)
    @app.get("/app/{path:path}", include_in_schema=False)
    async def spa_fallback(path: str | None = None):
        return FileResponse(WEB_DIST / "index.html")

    # Redirecionar raiz para /app/
    @app.get("/", include_in_schema=False)
    async def root_spa():
        return RedirectResponse(url="/app/")

    # Compat: redirecionar /web/* para /app/* (URLs documentadas)
    @app.get("/web", include_in_schema=False)
    async def web_root_redirect():
        return RedirectResponse(url="/app/")

    @app.get("/web/{path:path}", include_in_schema=False)
    async def web_prefix_redirect(path: str):
        # Ex.: /web/login -> /app/login
        return RedirectResponse(url=f"/app/{path}")

    # Alias legível para o Anexo II
    @app.get("/plano-atividades", include_in_schema=False)
    async def plano_atividades_alias_root():
        return RedirectResponse(url="/app/plano-atividades")

    @app.get("/plano-atividades/{path:path}", include_in_schema=False)
    async def plano_atividades_alias(path: str):
        return RedirectResponse(url=f"/app/plano-atividades/{path}")

# ------------------------------
# Admin: Substituir catálogos via CSV (CNES)
# ------------------------------
@app.post("/admin/catalogos/upload-cnes")
async def upload_catalogos_cnes(
    tipo: str = Form(...),  # 'instituicoes' ou 'unidades'
    file: UploadFile = File(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    user = auth.get_current_user(db, credentials.credentials)
    if user.tipo != 'admin':
        raise HTTPException(status_code=403, detail="Apenas administradores")

    if tipo not in ("instituicoes", "unidades"):
        raise HTTPException(status_code=400, detail="tipo deve ser 'instituicoes' ou 'unidades'")

    try:
        import csv, io
        content = await file.read()
        # Detectar encoding simples
        text = None
        for enc in ('utf-8', 'latin-1', 'iso-8859-1'):
            try:
                text = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise HTTPException(status_code=400, detail="Encoding do arquivo não suportado")

        # Detectar delimitador
        sample = text[:1024]
        delimiter = ';' if sample.count(';') > sample.count(',') else ','
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

        nomes = []
        for row in reader:
            # Procura colunas comuns: 'nome', 'NO_FANTASIA', 'NO_RAZAO_SOCIAL'
            nome = row.get('nome') or row.get('NO_FANTASIA') or row.get('NO_RAZAO_SOCIAL') or row.get('NM_UNIDADE')
            if nome:
                nome = str(nome).strip()
                if nome:
                    nomes.append(nome)

        if not nomes:
            raise HTTPException(status_code=400, detail="Nenhum nome encontrado no CSV")

        # Substituir catálogo selecionado
        if tipo == 'instituicoes':
            # Apagar todas e inserir novas
            existing = db.query(models.Instituicao).all()
            for e in existing:
                db.delete(e)
            db.commit()
            for n in nomes:
                db.add(models.Instituicao(nome=n))
            db.commit()
            return {"message": f"Instituições substituídas: {len(nomes)}"}
        else:
            existing = db.query(models.Unidade).all()
            for e in existing:
                db.delete(e)
            db.commit()
            for n in nomes:
                db.add(models.Unidade(nome=n))
            db.commit()
            return {"message": f"Unidades substituídas: {len(nomes)}"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar CSV: {e}")

@app.post("/admin/seed/unidades-cnes")
async def seed_unidades_cnes(
    dados: list[dict],
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    user = auth.get_current_user(db, credentials.credentials)
    if user.tipo != 'admin':
        raise HTTPException(status_code=403, detail="Apenas administradores")
    inseridos = 0
    atualizados = 0
    for item in dados:
        cnes = str(item.get('CNES') or item.get('cnes') or '').strip()
        nome = str(item.get('Nome Fantasia') or item.get('nome') or '').strip()
        razao = str(item.get('Razão Social') or item.get('razao_social') or '').strip()
        if not nome:
            continue
        # tenta achar por cnes ou por nome
        rec = None
        if cnes:
            rec = db.query(models.Unidade).filter(models.Unidade.cnes == cnes).first()
        if not rec:
            rec = db.query(models.Unidade).filter(models.Unidade.nome.ilike(nome)).first()
        if rec:
            rec.nome = nome or rec.nome
            if cnes:
                rec.cnes = cnes
            if razao:
                rec.razao_social = razao
            atualizados += 1
        else:
            db.add(models.Unidade(nome=nome, cnes=cnes or None, razao_social=razao or None))
            inseridos += 1
    db.commit()
    return {"message": f"Unidades processadas. Inseridos: {inseridos}, Atualizados: {atualizados}", "inseridos": inseridos, "atualizados": atualizados}

# ------------------------------
# Anexo II API (CRUD básico)
# ------------------------------
@app.get("/anexo2", response_model=List[schemas.Anexo2])
async def list_all_anexo2(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    user = auth.get_current_user(db, credentials.credentials)
    # Apenas admin pode listar todos
    if user.tipo != 'admin':
        raise HTTPException(status_code=403, detail="Apenas administradores")
    return crud.list_all_anexo2(db)
@app.get("/anexo2/{estagio_id}", response_model=schemas.Anexo2)
async def get_anexo2(estagio_id: int, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    auth.get_current_user(db, credentials.credentials)
    anexo = crud.get_anexo2_by_estagio(db, estagio_id)
    if not anexo:
        raise HTTPException(status_code=404, detail="Anexo II não encontrado")
    return anexo

@app.post("/anexo2", response_model=schemas.Anexo2)
async def create_anexo2(payload: schemas.Anexo2Create, label: Optional[str] = None, comment: Optional[str] = None, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    auth.get_current_user(db, credentials.credentials)
    # Se já existir, atualiza
    existing = crud.get_anexo2_by_estagio(db, payload.estagio_id)
    if existing:
        updated = crud.update_anexo2(db, existing.id, payload)
        # Salvar versão quando status final
        if (updated.status or "final") == "final":
            try:
                crud.save_anexo2_version(db, payload.estagio_id, updated, label=label, comment=comment)
            except Exception as e:
                print(f"⚠️ Falha ao salvar versão: {e}")
        return updated
    created = crud.create_anexo2(db, payload)
    if (created.status or "final") == "final":
        try:
            crud.save_anexo2_version(db, payload.estagio_id, created, label=label, comment=comment)
        except Exception as e:
            print(f"⚠️ Falha ao salvar versão: {e}")
    return created

@app.put("/anexo2/{anexo2_id}", response_model=schemas.Anexo2)
async def update_anexo2(anexo2_id: int, payload: schemas.Anexo2Create, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    auth.get_current_user(db, credentials.credentials)
    updated = crud.update_anexo2(db, anexo2_id, payload)
    if not updated:
        raise HTTPException(status_code=404, detail="Anexo II não encontrado")
    return updated

@app.get("/estagios/{estagio_id}/html", response_class=HTMLResponse)
async def get_estagio_html(estagio_id: int, request: Request, db: Session = Depends(get_db)):
    # Aceita token via query string (?token=) ou header Authorization: Bearer
    token = request.query_params.get('token')
    if not token:
        auth_header = request.headers.get('authorization') or request.headers.get('Authorization')
        if auth_header and auth_header.lower().startswith('bearer '):
            token = auth_header.split(' ', 1)[1].strip()
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    auth.get_current_user(db, token)
    e = crud.get_estagio_by_id(db, estagio_id)
    if not e:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")
    def _v(x): return x or '—'
    html = f"""
        <!doctype html>
        <html lang='pt-br'>
            <head>
                <meta charset='utf-8' />
                <meta name='viewport' content='width=device-width, initial-scale=1' />
                <title>Estágio #{e.id} - {_v(e.nome)}</title>
                <style>
                    body {{ font-family: -apple-system, BlinkMacSystemFont, Segoe UI, Roboto, Helvetica, Arial, sans-serif; color:#222; margin:24px; }}
                    h1 {{ margin: 0 0 16px 0; }}
                    .grid {{ display:grid; grid-template-columns: 1fr 1fr; gap:12px; }}
                    .card {{ background:#fff; border:1px solid #eee; border-radius:8px; padding:16px; margin-bottom:16px; }}
                    .label {{ font-size:12px; color:#666; }}
                    .value {{ font-size:15px; font-weight:600; }}
                    .muted {{ color:#666 }}
                    .header {{ display:flex; justify-content:space-between; align-items:center; }}
                    .badge {{ background:#e3f2fd; padding:4px 8px; border-radius:6px; font-size:12px; }}
                    @media print {{ .no-print {{ display:none }} body {{ margin:0 }} }}
                </style>
            </head>
            <body>
                <div class='header'>
                    <h1>Estágio #{e.id}</h1>
                    <div class='badge'>{_v(e.periodo)}</div>
                </div>
                <div class='card'>
                    <div class='grid'>
                        <div><div class='label'>Nome</div><div class='value'>{_v(e.nome)}</div></div>
                        <div><div class='label'>E-mail</div><div class='value'>{_v(e.email)}</div></div>
                        <div><div class='label'>Telefone</div><div class='value'>{_v(e.telefone)}</div></div>
                        <div><div class='label'>Curso</div><div class='value'>{_v(getattr(e.curso,'nome',None))}</div></div>
                        <div><div class='label'>Instituição</div><div class='value'>{_v(getattr(e.instituicao,'nome',None))}</div></div>
                        <div><div class='label'>Unidade</div><div class='value'>{_v(getattr(e.unidade,'nome',None))}</div></div>
                        <div><div class='label'>Supervisor</div><div class='value'>{_v(getattr(e.supervisor,'nome',None))}</div></div>
                        <div><div class='label'>Nível</div><div class='value'>{_v(e.nivel)}</div></div>
                        <div><div class='label'>Carga Horária</div><div class='value'>{_v(e.carga_horaria)}</div></div>
                        <div><div class='label'>Salário</div><div class='value'>{_v(e.salario)}</div></div>
                    </div>
                </div>
                <div class='card'>
                    <div class='label'>Disciplina</div>
                    <div class='value'>{_v(e.disciplina)}</div>
                </div>
                <div class='card'>
                    <div class='label'>Observações</div>
                    <div class='value'>{_v(e.observacoes)}</div>
                </div>
                <div class='muted'>Gerado em {__import__('datetime').datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
                <div class='no-print' style='margin-top:12px'>
                    <button onclick='window.print()'>Imprimir</button>
                </div>
            </body>
        </html>
        """
    return HTMLResponse(content=html)

@app.post("/anexo2/{source_estagio_id}/clone")
async def clonar_anexo_para_instituicoes(
    source_estagio_id: int,
    payload: dict = Body(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Clona o Anexo II de um estágio origem para múltiplas instituições de ensino.

    Body esperado:
      {
        "instituicao_ids": [1,2,3],    # opcional (se ausente e todas=true, usa todas)
        "todas": false,                 # opcional: clonar para todas as instituições do sistema (exceto a origem)
        "exercicio": "2025",          # opcional: sobrescrever exercício
        "status": "final",            # opcional: sobrescrever status (final|rascunho)
        "replace": false               # opcional: se já existir plano no destino, atualiza ao invés de pular
      }
    """
    auth.get_current_user(db, credentials.credentials)

    anexo_src = crud.get_anexo2_by_estagio(db, source_estagio_id)
    if not anexo_src:
        raise HTTPException(status_code=404, detail="Anexo II de origem não encontrado")
    estagio_src = crud.get_estagio_by_id(db, source_estagio_id)
    if not estagio_src:
        raise HTTPException(status_code=404, detail="Estágio de origem não encontrado")

    # Resolver instituições alvo
    ids = payload.get("instituicao_ids") or []
    todas = bool(payload.get("todas", False))
    if todas:
        insts = crud.get_instituicoes(db)
        ids = [i.id for i in insts if i.id != (estagio_src.instituicao_id or -1)]
    if not ids:
        raise HTTPException(status_code=400, detail="Nenhuma instituição alvo informada")

    override_exercicio = payload.get("exercicio") or anexo_src.exercicio
    override_status = payload.get("status") or anexo_src.status or "final"
    replace = bool(payload.get("replace", False))

    created = 0
    updated = 0
    skipped = 0
    results = []

    for inst_id in ids:
        inst = db.query(models.Instituicao).filter(models.Instituicao.id == inst_id).first()
        if not inst:
            results.append({"instituicao_id": inst_id, "status": "erro", "detail": "Instituição não encontrada"})
            continue

        # Encontrar/gerar curso_id compatível
        curso_id = estagio_src.curso_id
        # Encontrar/gerar Estágio destino (um por instituição+curso)
        estagio_dst = db.query(models.Estagio).filter(
            models.Estagio.instituicao_id == inst.id,
            models.Estagio.curso_id == curso_id
        ).first()
        if not estagio_dst:
            # Criar novo estágio sintético
            base_email = f"estagio.{inst.nome.lower().strip().replace(' ', '.')}@exemplo.com"
            email = base_email
            try:
                idx = 1
                while db.query(models.Estagio).filter(models.Estagio.email == email).first() is not None and idx < 100:
                    local, domain = base_email.split('@', 1)
                    email = f"{local}.{idx}@{domain}"
                    idx += 1
            except Exception:
                email = base_email
            estagio_dst = crud.create_estagio(db, schemas.EstagioCreate(
                nome=f"Estágio {inst.nome}",
                email=email,
                instituicao_id=inst.id,
                curso_id=curso_id,
                periodo=estagio_src.periodo or "",
                observacoes="Criado via clonagem do Anexo II"
            ))

        # Já existe plano no destino?
        existing_dst = crud.get_anexo2_by_estagio(db, estagio_dst.id)
        atividades = []
        for a in anexo_src.atividades:
            # Normalizar Unidade e Supervisor para catálogos (fuzzy match)
            unidade_normalizada = crud.normalize_unidade_nome(db, a.unidade_setor)
            supervisor_normalizado = crud.normalize_supervisor_nome(db, a.supervisor_nome)
            
            atividades.append(schemas.Anexo2AtividadeCreate(
                disciplina=a.disciplina,
                descricao=a.descricao,
                nivel=a.nivel,
                unidade_setor=unidade_normalizada or a.unidade_setor,
                data_inicio=a.data_inicio,
                data_fim=a.data_fim,
                horario=a.horario,
                dias_semana=a.dias_semana,
                quantidade_grupos=a.quantidade_grupos,
                num_estagiarios_por_grupo=a.num_estagiarios_por_grupo,
                carga_horaria_individual=a.carga_horaria_individual,
                supervisor_nome=supervisor_normalizado or a.supervisor_nome,
                supervisor_conselho=a.supervisor_conselho,
                valor=a.valor,
                territorio=getattr(a, 'territorio', None),
            ))

        anexo_payload = schemas.Anexo2Create(
            estagio_id=estagio_dst.id,
            instituicao_ensino=inst.nome,
            curso=anexo_src.curso,
            exercicio=override_exercicio,
            status=override_status,
            versao=1,
            logo_url=anexo_src.logo_url,
            atividades=atividades
        )

        if existing_dst:
            if replace:
                crud.update_anexo2(db, existing_dst.id, anexo_payload)
                updated += 1
                results.append({"instituicao_id": inst.id, "status": "updated", "estagio_id": estagio_dst.id})
            else:
                skipped += 1
                results.append({"instituicao_id": inst.id, "status": "skipped", "reason": "já existe plano"})
        else:
            crud.create_anexo2(db, anexo_payload)
            created += 1
            results.append({"instituicao_id": inst.id, "status": "created", "estagio_id": estagio_dst.id})

    return {"created": created, "updated": updated, "skipped": skipped, "results": results}

@app.get("/anexo2/{estagio_id}/versions", response_model=List[schemas.Anexo2Version])
async def list_versions(estagio_id: int, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    auth.get_current_user(db, credentials.credentials)
    return crud.list_anexo2_versions(db, estagio_id)

@app.post("/anexo2/{estagio_id}/versions/{versao}/restore", response_model=schemas.Anexo2)
async def restore_version(estagio_id: int, versao: int, credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    auth.get_current_user(db, credentials.credentials)
    restored = crud.restore_anexo2_version(db, estagio_id, versao)
    if not restored:
        raise HTTPException(status_code=404, detail="Versão não encontrada")
    return restored

# Endpoint específico para cadastro manual de atividades
@app.post("/anexo2/cadastro-manual", response_model=schemas.Anexo2)
async def create_anexo2_manual(
    request: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security), 
    db: Session = Depends(get_db)
):
    """Endpoint para criar Anexo II através do cadastro manual de atividades"""
    auth.get_current_user(db, credentials.credentials)
    
    estagio_id = request.get("estagio_id")
    atividades_data = request.get("atividades", [])
    
    if not estagio_id:
        raise HTTPException(status_code=400, detail="estagio_id é obrigatório")
    
    if not atividades_data:
        raise HTTPException(status_code=400, detail="Pelo menos uma atividade é obrigatória")
    
    # Verificar se o estágio existe
    estagio = crud.get_estagio(db, estagio_id)
    if not estagio:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")
    
    # Converter atividades para o formato do esquema
    atividades_converted = []
    for ativ in atividades_data:
        atividade = schemas.Anexo2AtividadeCreate(
            disciplina=ativ.get("disciplina"),
            descricao=f"Disciplina: {ativ.get('disciplina', '')} | Supervisor: {ativ.get('supervisor', '')}",
            nivel=ativ.get("nivel"),
            unidade_setor=ativ.get("instituicao"),
            data_inicio=ativ.get("data_inicio"),
            data_fim=ativ.get("data_fim"),
            horario=ativ.get("horario"),
            dias_semana=ativ.get("periodo"),
            quantidade_grupos=ativ.get("grupos", 1),
            num_estagiarios_por_grupo=ativ.get("num_estagiarios", 1),
            carga_horaria_individual=str(ativ.get("horas_calculadas", 0)),
            supervisor_nome=ativ.get("supervisor"),
            supervisor_conselho="",
            valor=str(ativ.get("horas_calculadas", 0))
        )
        atividades_converted.append(atividade)
    
    # Criar o payload do Anexo II
    anexo2_data = schemas.Anexo2Create(
        estagio_id=estagio_id,
        instituicao_ensino=estagio.instituicao.nome if estagio.instituicao else "",
        curso=estagio.curso.nome if estagio.curso else "",
        exercicio="2024",  # Ano atual, pode ser parametrizado
        status="final",
        versao=1,
        atividades=atividades_converted
    )
    
    # Verificar se já existe anexo para este estágio
    existing = crud.get_anexo2_by_estagio(db, estagio_id)
    if existing:
        # Atualizar existente
        updated = crud.update_anexo2(db, existing.id, anexo2_data)
        try:
            crud.save_anexo2_version(db, estagio_id, updated, label="Cadastro Manual", comment="Criado via cadastro manual")
        except Exception as e:
            print(f"⚠️ Falha ao salvar versão: {e}")
        return updated
    else:
        # Criar novo
        created = crud.create_anexo2(db, anexo2_data)
        try:
            crud.save_anexo2_version(db, estagio_id, created, label="Cadastro Manual", comment="Criado via cadastro manual")
        except Exception as e:
            print(f"⚠️ Falha ao salvar versão: {e}")
        return created

# ------------------------------
# Planos Anexo II - busca paginada
# ------------------------------
@app.get("/planos/search")
async def search_planos(
    curso: Optional[str] = None,
    exercicio: Optional[str] = None,
    instituicao: Optional[str] = None,
    limit: int = 20,
    offset: int = 0,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    user = auth.get_current_user(db, credentials.credentials)
    limit = max(1, min(200, limit))
    if user.tipo == 'admin':
        data = crud.search_planos_anexo2(db, curso=curso, exercicio=exercicio, instituicao=instituicao, skip=offset, limit=limit)
    elif user.tipo == 'supervisor':
        # Se usuário for supervisor, tentar localizar supervisor vinculado (email match) para filtrar
        supervisor = db.query(models.Supervisor).filter(models.Supervisor.email == user.email).first()
        if supervisor:
            data = crud.list_planos_for_supervisor(db, supervisor_id=supervisor.id, curso=curso, exercicio=exercicio, instituicao=instituicao, skip=offset, limit=limit)
        else:
            data = {"total": 0, "items": []}
    elif user.tipo == 'escola':
        if getattr(user, 'instituicao_id', None):
            data = crud.list_planos_for_instituicao(db, instituicao_id=user.instituicao_id, curso=curso, exercicio=exercicio, instituicao=instituicao, skip=offset, limit=limit)
        else:
            data = {"total": 0, "items": []}
    else:
        data = {"total": 0, "items": []}
    # Serializar manualmente porque response_model dinâmico
    payload = {
        "total": data["total"],
        "items": [schemas.Anexo2.model_validate(p) for p in data["items"]]
    }
    # Garantir serialização segura (datetime, etc.)
    return JSONResponse(content=jsonable_encoder(payload))
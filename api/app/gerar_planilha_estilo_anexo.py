#!/usr/bin/env python3
"""
Gera uma planilha Excel (.xlsx) no estilo visual do anexo exibido (com faixas coloridas)
contendo blocos de linhas por unidade. Cada bloco tem cabeçalho com:
Instituição | CURSO | NÍVEL | INÍCIO | FIM | HORÁRIO | PERÍODO | SUPERVISOR | DISCIPLINA | Nº de estagiários

Fonte de dados: plano base de um estágio (atividades) + lista de unidades. As atividades
são replicadas (ou placeholders) para cada unidade.

Uso:
  python gerar_planilha_estilo_anexo.py --api-url http://localhost:8000 --token TOKEN \
      --estagio-base 4 --arquivo saida.xlsx --placeholder

  # Com datas sequenciais por unidade iniciando em 22/07/2025
  python gerar_planilha_estilo_anexo.py --estagio-base 4 --data-inicio 22/07/2025 --sequencial-semanas

"""
import os, sys, argparse, requests
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

HEADERS = ["Instituição","CURSO","NÍVEL","INÍCIO","FIM","HORÁRIO","PERÍODO","SUPERVISOR","DISCIPLINA","Nº de estagiários"]

TEAL = PatternFill(start_color="008B9DA4", end_color="008B9DA4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style='thin', color='000000')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fetch_json(url, headers):
    r = requests.get(url, headers=headers)
    if not r.ok:
        raise RuntimeError(f"GET {url} -> {r.status_code} {r.text}")
    return r.json()


def gerar_datas(base: str, n: int):
    try:
        d0 = datetime.strptime(base, '%d/%m/%Y')
    except:
        return [("","")]*n
    out=[]
    for i in range(n):
        ini = d0 + timedelta(weeks=i)
        fim = ini + timedelta(days= (4 if i==0 else 4))
        out.append((ini.strftime('%d/%m/%Y'), fim.strftime('%d/%m/%Y')))
    return out


def ajustar_larguras(ws):
    widths = [18,16,8,12,12,22,16,28,26,16]
    for col, w in zip(range(1,len(widths)+1), widths):
        ws.column_dimensions[chr(64+col)].width = w


def escrever_bloco(ws, start_row, unidade_nome, atividades, instituicao, curso, datas_seq=None, placeholder=False):
    # Linha de faixa colorida com nome da unidade centralizado entre as colunas A:J
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
    cell = ws.cell(row=start_row, column=1, value=unidade_nome.upper())
    cell.fill = TEAL
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.alignment = CENTER
    cell.border = BORDER
    for c in range(2,11):
        ws.cell(row=start_row, column=c).border = BORDER

    # Cabeçalho
    header_row = start_row + 1
    for idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=h)
        cell.fill = TEAL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    row = header_row + 1
    for i, at in enumerate(atividades):
        if datas_seq and i < len(datas_seq):
            di, df = datas_seq[i]
        else:
            di = at.get('data_inicio','')
            df = at.get('data_fim','')
        if placeholder:
            di = di or '[DD/MM/AAAA]'
            df = df or '[DD/MM/AAAA]'
        values = [
            instituicao,
            curso,
            at.get('nivel',''),
            di,
            df,
            at.get('horario',''),
            at.get('periodo','') or at.get('PERÍODO','') or '',
            at.get('supervisor_nome',''),
            at.get('disciplina', at.get('DISCIPLINA','')),
            at.get('num_estagiarios_por_grupo', at.get('Nº de estagiários','')) or 1
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = LEFT if col in (1,2,8,9) else CENTER
            c.border = BORDER
        row += 1

    return row + 1  # blank line after bloco


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--api-url', default='http://localhost:8000')
    ap.add_argument('--token')
    ap.add_argument('--estagio-base', type=int, required=True)
    ap.add_argument('--arquivo', default='planilha_anexo.xlsx')
    ap.add_argument('--data-inicio', help='Data início para datas sequenciais')
    ap.add_argument('--sequencial-semanas', action='store_true', help='Distribui atividades em semanas consecutivas')
    ap.add_argument('--placeholder', action='store_true', help='Usa placeholders em datas vazias')
    ap.add_argument('--limitar-atividades', type=int, help='Limita número de atividades usadas do plano base')
    args = ap.parse_args()

    token = args.token
    if not token:
        token_file = os.path.join(os.path.dirname(__file__), '.token')
        if os.path.exists(token_file):
            token = open(token_file).read().strip()
    if not token:
        print('Necessário --token ou arquivo .token')
        return 1
    headers={'Authorization': f'Bearer {token}'}

    try:
        plano = fetch_json(f"{args.api_url}/anexo2/{args.estagio_base}", headers)
    except Exception as e:
        print(f'Falha ao obter plano base: {e}')
        return 1
    atividades = plano.get('atividades', [])
    if args.limitar_atividades:
        atividades = atividades[:args.limitar_atividades]

    try:
        unidades = fetch_json(f"{args.api_url}/unidades", headers)
    except Exception as e:
        print(f'Falha ao listar unidades: {e}')
        return 1

    instituicao = plano.get('instituicao_ensino','Instituição')
    curso = plano.get('curso','Curso')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Anexo'
    ajustar_larguras(ws)

    datas_seq = None
    if args.sequencial_semanas and args.data_inicio:
        datas_seq = gerar_datas(args.data_inicio, len(atividades))

    row = 1
    for u in unidades:
        row = escrever_bloco(
            ws, row, u['nome'], atividades, instituicao, curso,
            datas_seq=datas_seq, placeholder=args.placeholder
        )

    wb.save(args.arquivo)
    print(f'Arquivo gerado: {os.path.abspath(args.arquivo)}')
    return 0

if __name__ == '__main__':
    sys.exit(main())

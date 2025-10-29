#!/usr/bin/env python3
"""
Script para analisar planilhas Excel existentes e identificar:
1. Estrutura de colunas
2. Fórmulas de cálculo de horas
3. Padrões de dados
4. Possíveis melhorias no importador
"""
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl não encontrado. Instale com: pip install openpyxl")
    sys.exit(1)

def analisar_planilha(arquivo_path):
    print(f"\n{'='*60}")
    print(f"📋 ANALISANDO: {os.path.basename(arquivo_path)}")
    print(f"{'='*60}")
    
    try:
        wb = openpyxl.load_workbook(arquivo_path, data_only=False)  # Preservar fórmulas
        
        print(f"📊 Número de abas: {len(wb.sheetnames)}")
        print(f"📝 Abas disponíveis: {', '.join(wb.sheetnames)}")
        
        for i, sheet_name in enumerate(wb.sheetnames):
            print(f"\n--- ABA {i+1}: {sheet_name} ---")
            ws = wb[sheet_name]
            
            print(f"📐 Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")
            
            # Analisar primeiras 5 linhas para encontrar cabeçalho
            print("\n🔍 Primeiras 5 linhas:")
            for row in range(1, min(6, ws.max_row + 1)):
                valores = []
                for col in range(1, min(ws.max_column + 1, 15)):  # Máx 15 colunas
                    cell = ws.cell(row=row, column=col)
                    valor = cell.value
                    if valor is not None:
                        # Mostrar fórmula se existir
                        if str(valor).startswith('='):
                            valores.append(f"[FÓRMULA] {valor}")
                        else:
                            valores.append(str(valor)[:30])  # Truncar valores longos
                    else:
                        valores.append("")
                print(f"Linha {row}: {' | '.join(valores[:10])}")  # Mostrar até 10 colunas
            
            # Procurar por fórmulas (especialmente relacionadas a horas/cálculos)
            print("\n🧮 Procurando fórmulas de cálculo:")
            formulas_encontradas = []
            for row in range(1, min(ws.max_row + 1, 100)):  # Verificar até 100 linhas
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and str(cell.value).startswith('='):
                        formula = str(cell.value)
                        if any(term in formula.upper() for term in ['SUM', 'SOMA', 'HOUR', 'TIME', 'DATA', '*', '+']):
                            coord = f"{get_column_letter(col)}{row}"
                            formulas_encontradas.append(f"{coord}: {formula}")
            
            if formulas_encontradas:
                print("Fórmulas relevantes encontradas:")
                for formula in formulas_encontradas[:10]:  # Mostrar até 10
                    print(f"  • {formula}")
                if len(formulas_encontradas) > 10:
                    print(f"  ... e mais {len(formulas_encontradas) - 10} fórmulas")
            else:
                print("Nenhuma fórmula de cálculo identificada.")
            
            # Identificar possível linha de cabeçalho
            print("\n📋 Identificando cabeçalho provável:")
            palavras_chave = [
                'disciplina', 'descricao', 'nivel', 'unidade', 'setor', 'data', 'inicio', 'fim',
                'horario', 'dias', 'semana', 'grupos', 'estagiarios', 'carga', 'horaria', 
                'supervisor', 'conselho', 'valor', 'atividade'
            ]
            
            melhor_linha = 1
            melhor_score = 0
            
            for row in range(1, min(11, ws.max_row + 1)):  # Verificar até linha 10
                score = 0
                valores_linha = []
                for col in range(1, ws.max_column + 1):
                    valor = ws.cell(row=row, column=col).value
                    if valor:
                        valor_str = str(valor).lower()
                        valores_linha.append(valor_str[:20])
                        for palavra in palavras_chave:
                            if palavra in valor_str:
                                score += 1
                
                print(f"Linha {row} (score: {score}): {' | '.join(valores_linha[:8])}")
                if score > melhor_score:
                    melhor_score = score
                    melhor_linha = row
            
            print(f"\n✅ Linha de cabeçalho mais provável: {melhor_linha} (score: {melhor_score})")
    
    except Exception as e:
        print(f"❌ Erro ao analisar {arquivo_path}: {e}")

def main():
    pasta_app = Path(__file__).parent
    
    # Procurar por arquivos Excel
    arquivos_excel = list(pasta_app.glob("*.xlsx")) + list(pasta_app.glob("*.xls"))
    
    if not arquivos_excel:
        print("❌ Nenhum arquivo Excel encontrado na pasta atual.")
        return
    
    print(f"🔍 Encontrados {len(arquivos_excel)} arquivo(s) Excel:")
    for arquivo in arquivos_excel:
        print(f"  • {arquivo.name}")
    
    # Analisar cada arquivo
    for arquivo in arquivos_excel:
        analisar_planilha(arquivo)
    
    print(f"\n{'='*60}")
    print("📋 RESUMO DA ANÁLISE")
    print(f"{'='*60}")
    print("Com base na análise das planilhas, posso:")
    print("1. ✅ Adaptar o importador para detectar automaticamente o cabeçalho")
    print("2. ✅ Preservar/replicar fórmulas de cálculo de horas no backend")
    print("3. ✅ Criar templates específicos baseados nessas estruturas")
    print("4. ✅ Melhorar a normalização de dados (datas, horários, etc.)")
    print("\nPróximo passo: Implementar as melhorias identificadas!")

if __name__ == "__main__":
    main()
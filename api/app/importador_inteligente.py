"""
Importador Inteligente de Planilhas Excel
Detecta automaticamente a estrutura e extrai dados das planilhas de estágio
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional, Tuple
import re
from datetime import datetime
import unicodedata

class ImportadorInteligente:
    
    # Padrões conhecidos baseados na análise das planilhas
    HEADERS_CONHECIDOS = [
        'INSTITUIÇÃO DE ENSINO',
        'CURSO',
        'NÍVEL',
        'INÍCIO',
        'FIM', 
        'HORÁRIO',
        'PERÍODO',
        'DIAS DA SEMANA',
        'SUPERVISOR',
        'DISCIPLINA',
        'DESCRIÇÃO',
        'Nº CONSELHO',
        'VALOR',
        'UNIDADE/ SETOR',
        'Nº de estagiários'
    ]
    
    NIVEIS_VALIDOS = ['G', 'I', 'PG', 'M']
    
    def __init__(self):
        self.dados_detectados = []
        self.estrutura = {}
        self.erros = []
        
    def analisar_planilha(self, caminho_arquivo: str) -> Dict[str, Any]:
        """Analisa uma planilha Excel e detecta sua estrutura automaticamente"""
        
        try:
            # Carregar workbook
            wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
            
            resultado = {
                'arquivo': caminho_arquivo,
                'abas': [],
                'total_atividades': 0,
                'estrutura_detectada': True,
                'erros': []
            }
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Analisar cada aba
                aba_info = self._analisar_aba(sheet, sheet_name)
                if aba_info['atividades']:
                    resultado['abas'].append(aba_info)
                    resultado['total_atividades'] += len(aba_info['atividades'])
            
            wb.close()
            return resultado
            
        except Exception as e:
            return {
                'arquivo': caminho_arquivo,
                'abas': [],
                'total_atividades': 0,
                'estrutura_detectada': False,
                'erros': [f"Erro ao analisar arquivo: {str(e)}"]
            }
    
    def _analisar_aba(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Analisa uma aba específica da planilha"""

        aba_info = {
            'nome': sheet_name,
            'linha_cabecalho': None,
            'colunas_detectadas': {},
            'atividades': [],
            'total_horas_calculadas': 0,
            'candidatos_cabecalho': []
        }

        # Procurar linha do cabeçalho (geralmente linha 6 baseado na análise)
        linha_cabecalho, cand = self._detectar_linha_cabecalho(sheet, return_candidates=True)
        aba_info['candidatos_cabecalho'] = cand

        if linha_cabecalho is None:
            return aba_info

        aba_info['linha_cabecalho'] = linha_cabecalho

        # Mapear colunas
        colunas = self._mapear_colunas(sheet, linha_cabecalho)
        aba_info['colunas_detectadas'] = colunas

        # Extrair dados
        atividades = self._extrair_atividades(sheet, linha_cabecalho + 1, colunas)
        aba_info['atividades'] = atividades

        # Calcular total de horas
        total_horas = sum(ativ.get('horas_calculadas', 0) for ativ in atividades)
        aba_info['total_horas_calculadas'] = total_horas

        return aba_info
    
    def _detectar_linha_cabecalho(self, sheet, return_candidates: bool = False):
        """Detecta a linha onde está o cabeçalho. Retorna também candidatos para debug quando solicitado."""
        candidatos: List[Dict[str, Any]] = []
        linha_escolhida = None

        for row_num in range(1, min(sheet.max_row, 40) + 1):  # ampliar busca
            try:
                row = sheet[row_num]
            except Exception:
                continue
            headers_encontrados = 0
            valores = []
            for cell in row:
                val = cell.value
                if val is not None:
                    valores.append(str(val))
                if val and isinstance(val, str):
                    texto = val.strip().upper()
                    for header in self.HEADERS_CONHECIDOS:
                        if header.upper() in texto:
                            headers_encontrados += 1
                            break
            candidatos.append({
                'linha': row_num,
                'match_headers': headers_encontrados,
                'valores_primeiras_colunas': valores[:10]
            })
            if headers_encontrados >= 4 and linha_escolhida is None:
                linha_escolhida = row_num

        if return_candidates:
            return linha_escolhida, candidatos
        return linha_escolhida
    
    def _mapear_colunas(self, sheet, linha_cabecalho: int) -> Dict[str, int]:
        """Mapeia as colunas baseado no cabeçalho detectado"""
        
        mapeamento = {}
        row = sheet[linha_cabecalho]
        
        for col_num, cell in enumerate(row, 1):
            if cell.value and isinstance(cell.value, str):
                texto = cell.value.strip().upper()
                
                # Mapear para campos padronizados
                if ('UNIDADE' in texto and 'SETOR' in texto) or 'UNIDADE/ SETOR' in texto or 'UNIDADE / SETOR' in texto:
                    mapeamento['unidade_setor'] = col_num
                elif 'INSTITUIÇÃO' in texto or 'INSTITUICAO' in texto:
                    mapeamento['instituicao'] = col_num
                elif 'CURSO' in texto:
                    mapeamento['curso'] = col_num
                elif 'NÍVEL' in texto or 'NIVEL' in texto:
                    mapeamento['nivel'] = col_num
                elif 'INÍCIO' in texto or 'INICIO' in texto:
                    mapeamento['data_inicio'] = col_num
                elif 'FIM' in texto:
                    mapeamento['data_fim'] = col_num
                elif 'HORÁRIO' in texto or 'HORARIO' in texto:
                    mapeamento['horario'] = col_num
                elif 'PERÍODO' in texto or 'PERIODO' in texto:
                    mapeamento['periodo'] = col_num
                elif 'DIAS' in texto and 'SEMANA' in texto:
                    mapeamento['dias_semana'] = col_num
                elif 'SUPERVISOR' in texto:
                    mapeamento['supervisor'] = col_num
                elif 'DISCIPLINA' in texto:
                    mapeamento['disciplina'] = col_num
                elif ('ESTAGI' in texto and 'GRUPO' in texto) or 'Nº DE ESTAGIÁRIOS POR GRUPO' in texto or 'N DE ESTAGIARIOS POR GRUPO' in texto:
                    mapeamento['num_estagiarios'] = col_num
                elif 'ESTAGIÁRIOS' in texto or 'ESTAGIARIOS' in texto:
                    mapeamento['num_estagiarios'] = col_num
                elif 'DESCRI' in texto:  # descrição de atividades
                    mapeamento['descricao'] = col_num
                elif 'CONSELHO' in texto:
                    mapeamento['supervisor_conselho'] = col_num
                elif texto.strip() == 'VALOR' or ' VALOR' in texto or texto.startswith('VALOR'):
                    mapeamento['valor'] = col_num
                elif 'GRUPOS' in texto:
                    mapeamento['grupos'] = col_num
                elif ('HORAS' in texto and 'INDIVIDUAIS' in texto) or ('CARGA' in texto and 'INDIVIDUAL' in texto):
                    mapeamento['horas_individuais'] = col_num
                    
        return mapeamento
    
    def _extrair_atividades(self, sheet, linha_inicio: int, colunas: Dict[str, int]) -> List[Dict[str, Any]]:
        """Extrai as atividades da planilha com forward-fill e tolerância a células vazias."""

        atividades: List[Dict[str, Any]] = []
        last_instituicao = None
        last_curso = None
        last_nivel = None

        for row_num in range(linha_inicio, sheet.max_row + 1):
            try:
                row = sheet[row_num]
            except Exception:
                continue

            atividade: Dict[str, Any] = {}

            for campo, col_num in colunas.items():
                if col_num <= len(row):
                    cell = row[col_num - 1]
                    valor = cell.value
                    if valor is not None and str(valor).strip() != '':
                        if campo in ['data_inicio', 'data_fim']:
                            atividade[campo] = self._converter_data(valor)
                        elif campo in ['num_estagiarios', 'grupos']:
                            atividade[campo] = self._converter_inteiro(valor)
                        elif campo == 'horas_individuais':
                            atividade[campo] = self._converter_float(valor)
                        elif campo == 'horario':
                            atividade[campo] = self._normalizar_horario(str(valor))
                        elif campo == 'dias_semana':
                            atividade[campo] = self._normalizar_dias_semana(str(valor))
                        else:
                            atividade[campo] = str(valor).strip()

            # Forward fill para instituicao / curso / nivel se ausentes nesta linha
            if 'instituicao' in colunas:
                if atividade.get('instituicao'):
                    last_instituicao = atividade['instituicao']
                elif last_instituicao:
                    atividade['instituicao'] = last_instituicao
            if 'curso' in colunas:
                if atividade.get('curso'):
                    last_curso = atividade['curso']
                elif last_curso:
                    atividade['curso'] = last_curso
            if 'nivel' in colunas:
                if atividade.get('nivel'):
                    last_nivel = atividade['nivel']
                elif last_nivel:
                    atividade['nivel'] = last_nivel

            # Pular linhas totalmente vazias
            if not any(atividade.values()):
                continue

            if self._atividade_valida(atividade):
                atividade['horas_calculadas'] = self._calcular_horas(atividade)
                atividades.append(atividade)

        return atividades
    
    def _atividade_valida(self, atividade: Dict[str, Any]) -> bool:
        """Verifica se uma atividade tem dados mínimos. Relaxado para cenários com mesclagem de células."""
        # Pelo menos dois entre curso, disciplina, supervisor ou instituicao
        chaves = ['curso', 'disciplina', 'supervisor', 'instituicao']
        preenchidos = sum(1 for k in chaves if atividade.get(k))
        return preenchidos >= 2
    
    def _calcular_horas(self, atividade: Dict[str, Any]) -> int:
        """Calcula as horas usando a fórmula descoberta nas planilhas"""
        
        nivel = atividade.get('nivel', 'G')
        grupos = atividade.get('grupos', 1)
        num_estagiarios = atividade.get('num_estagiarios', 1)
        horas_individuais = atividade.get('horas_individuais', 4)
        
        # Fórmula: =IFS(C9="G",I9*J9*K9*2,C9="I",I9*J9*K9*2,C9="PG",I9*J9*K9*2,C9="M",I9*J9*K9*1)
        multiplicador = 1 if nivel == 'M' else 2
        
        try:
            return int(grupos * num_estagiarios * horas_individuais * multiplicador)
        except:
            return 0
    
    def _converter_data(self, valor) -> str:
        """Converte valor para string de data"""
        if isinstance(valor, datetime):
            return valor.strftime('%Y-%m-%d')
        s = str(valor).strip()
        # Tenta formatos comuns: dd/mm/yyyy, dd-mm-yyyy
        m = re.match(r'^(\d{1,2})[\/-](\d{1,2})[\/-](\d{2,4})$', s)
        if m:
            d, mth, y = m.groups()
            if len(y) == 2:
                y = '20' + y  # heurística simples
            try:
                dt = datetime(int(y), int(mth), int(d))
                return dt.strftime('%Y-%m-%d')
            except Exception:
                return s
        return s
    
    def _converter_inteiro(self, valor) -> int:
        """Converte valor para inteiro"""
        try:
            if isinstance(valor, (int, float)):
                return int(valor)
            return int(float(str(valor).replace(',', '.')))
        except:
            return 1
    
    def _converter_float(self, valor) -> float:
        """Converte valor para float"""
        try:
            if isinstance(valor, (int, float)):
                return float(valor)
            return float(str(valor).replace(',', '.'))
        except:
            return 4.0

    # ========================
    # Normalizações auxiliares
    # ========================
    def _strip_accents(self, s: str) -> str:
        try:
            nfkd = unicodedata.normalize('NFD', s)
            return ''.join(c for c in nfkd if not unicodedata.category(c).startswith('M'))
        except Exception:
            return s

    def _normalizar_horario(self, s: str) -> str:
        """Normaliza variados formatos para 'HH:MM às HH:MM'"""
        if not s:
            return ''
        # Substitui 'as' por 'às' padrão visual apenas; extração usa dígitos
        s_clean = s.replace(' as ', ' às ').replace('AS', 'às')
        # Extrai números (horas e minutos)
        nums = re.findall(r'\d+', s_clean)
        # Espera 3-4 números (H,M,H,M). Se 3, assume minutos faltantes como 00 para o segundo horário
        if len(nums) < 2:
            return s.strip()
        # Monta pares
        h1 = nums[0].zfill(2)[:2]
        m1 = (nums[1] if len(nums) > 1 else '00').zfill(2)[:2]
        h2 = (nums[2] if len(nums) > 2 else nums[0]).zfill(2)[:2]
        m2 = (nums[3] if len(nums) > 3 else (nums[1] if len(nums) > 1 else '00')).zfill(2)[:2]
        # Clamps simples
        def clamp(h, m):
            try:
                hh = max(0, min(23, int(h)))
                mm = max(0, min(59, int(m)))
            except Exception:
                hh, mm = 0, 0
            return f"{hh:02d}:{mm:02d}"
        s1 = clamp(h1, m1)
        s2 = clamp(h2, m2)
        return f"{s1} às {s2}"

    def _normalizar_dias_semana(self, s: str) -> str:
        """Normaliza variações de dia da semana para forma curta: Seg, Ter, Qua, Qui, Sex, Sáb, Dom"""
        if not s:
            return ''
        base = self._strip_accents(s).lower().strip()
        base = base.replace('-feira', '').replace(' feira', '')
        mapa = {
            'segunda': 'Seg', 'seg': 'Seg', 'segunda-feira': 'Seg',
            'terca': 'Ter', 'terca-feira': 'Ter', 'ter': 'Ter',
            'quarta': 'Qua', 'quarta-feira': 'Qua', 'qua': 'Qua',
            'quinta': 'Qui', 'quinta-feira': 'Qui', 'qui': 'Qui',
            'sexta': 'Sex', 'sexta-feira': 'Sex', 'sex': 'Sex',
            'sabado': 'Sáb', 'sab': 'Sáb',
            'domingo': 'Dom', 'dom': 'Dom',
        }
        # tenta match direto
        if base in mapa:
            return mapa[base]
        # tenta por primeira palavra
        p = base.split()[0]
        return mapa.get(p, s.strip())
    
    def processar_multiplas_planilhas(self, caminhos: List[str]) -> Dict[str, Any]:
        """Processa múltiplas planilhas de uma vez"""
        
        resultado = {
            'arquivos_processados': 0,
            'total_atividades': 0,
            'planilhas': [],
            'resumo_por_curso': {},
            'erros_gerais': []
        }
        
        for caminho in caminhos:
            try:
                analise = self.analisar_planilha(caminho)
                resultado['planilhas'].append(analise)
                resultado['arquivos_processados'] += 1
                resultado['total_atividades'] += analise['total_atividades']
                
                # Resumo por curso
                for aba in analise['abas']:
                    for atividade in aba['atividades']:
                        curso = atividade.get('curso', 'Não identificado')
                        if curso not in resultado['resumo_por_curso']:
                            resultado['resumo_por_curso'][curso] = {
                                'atividades': 0,
                                'horas_totais': 0,
                                'instituicoes': set()
                            }
                        
                        resultado['resumo_por_curso'][curso]['atividades'] += 1
                        resultado['resumo_por_curso'][curso]['horas_totais'] += atividade.get('horas_calculadas', 0)
                        resultado['resumo_por_curso'][curso]['instituicoes'].add(atividade.get('instituicao', ''))
                        
            except Exception as e:
                resultado['erros_gerais'].append(f"Erro ao processar {caminho}: {str(e)}")
        
        # Converter sets para listas para serialização
        for curso_info in resultado['resumo_por_curso'].values():
            curso_info['instituicoes'] = list(curso_info['instituicoes'])
            
        return resultado
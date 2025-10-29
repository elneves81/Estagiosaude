#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importa uma planilha Excel (ou CSV) diretamente para o Anexo II (Plano de Atividades) de um estágio.

Uso básico:
  python importar_anexo2_from_excel.py --arquivo "COLÉGIO ESTADUAL ANA VANDA BASSARA.xlsx" --estagio-id 12 --status final \
      --instituicao "Colégio Estadual Ana Vanda Bassara" --curso "Enfermagem" --exercicio 2025 \
      --label "Importação inicial" --comment "Gerado via script"

Descobrir estágios:
  python importar_anexo2_from_excel.py --listar-estagios

Filtrar estágios por nome (case-insensitive substring):
  python importar_anexo2_from_excel.py --listar-estagios --filtro Bassara

Se o Anexo II já existir para o estágio, será substituído (atividades antigas removidas).

Heurística de detecção de cabeçalho e mapeamento de colunas replicada de /importar/preview.
"""
from __future__ import annotations
import argparse
import os
import sys
import json
import unicodedata
from typing import List, Dict, Any

# Ajuste de path: subir dois níveis para garantir import do pacote api.app
CURRENT_DIR = os.path.abspath(os.path.dirname(__file__))
API_DIR = os.path.abspath(os.path.join(CURRENT_DIR, '..'))
ROOT_DIR = os.path.abspath(os.path.join(API_DIR, '..'))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)
if API_DIR not in sys.path:
    sys.path.insert(0, API_DIR)

try:
    from api.app import db, crud, models, schemas  # quando executado a partir da raiz do projeto
except ModuleNotFoundError:
    try:
        from app import db, crud, models, schemas  # fallback quando corrente já é api/app
    except ModuleNotFoundError as e:
        print("❌ Não foi possível importar módulos do pacote app.")
        print("Paths:", sys.path[:5])
        raise e

from sqlalchemy.orm import Session

# ---------- Utilidades de normalização (reaproveitadas do frontend) ----------
DIAS_ALIAS = {
    'seg': 'Seg','segunda': 'Seg','segunda-feira': 'Seg',
    'ter': 'Ter','terça': 'Ter','terca': 'Ter','terça-feira': 'Ter','terca-feira': 'Ter',
    'qua': 'Qua','quarta': 'Qua','quarta-feira': 'Qua',
    'qui': 'Qui','quinta': 'Qui','quinta-feira': 'Qui',
    'sex': 'Sex','sexta': 'Sex','sexta-feira': 'Sex',
    'sab': 'Sáb','sáb': 'Sáb','sabado': 'Sáb','sábado': 'Sáb',
    'dom': 'Dom','domingo': 'Dom'
}
DIAS_ORD = ['Seg','Ter','Qua','Qui','Sex','Sáb','Dom']

def ordenar_dias(dias: List[str]) -> str:
    return ', '.join([d for d in DIAS_ORD if d in dias])

def parse_dias(txt: str) -> str:
    if not txt:
        return ''
    s = unicodedata.normalize('NFD', txt.lower())
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    import re
    # Ranges tipo 'seg a sex'
    m = re.search(r'(seg|ter|qua|qui|sex|sab|dom)\s*(?:a|-|ate)\s*(seg|ter|qua|qui|sex|sab|dom)', s)
    if m:
        seq = ['seg','ter','qua','qui','sex','sab','dom']
        i1, i2 = seq.index(m.group(1)), seq.index(m.group(2))
        if i1 <= i2:
            dias = [DIAS_ALIAS.get(x, '') for x in seq[i1:i2+1]]
            dias = [d for d in dias if d]
            return ordenar_dias(list(dict.fromkeys(dias)))
    parts = [p for p in re.split(r'[\s,\/;]+', s) if p]
    dias = [DIAS_ALIAS.get(p, '') for p in parts if DIAS_ALIAS.get(p, '')]
    return ordenar_dias(list(dict.fromkeys(dias)))

def normalizar_horario(s: str) -> str:
    if not s:
        return ''
    import re
    # Corrigir ' as ' sem acento para ' às '
    s = re.sub(r'\b(as)\b', 'às', s, flags=re.IGNORECASE)
    nums = re.findall(r'\d+', s)
    nums = [n.zfill(2) for n in nums]
    if len(nums) < 2:
        return s.strip()
    h1, m1 = nums[0][:2], (nums[1] if len(nums[1]) == 2 else '00')
    h2 = (nums[2] if len(nums) > 2 else nums[1])[:2]
    m2 = (nums[3] if len(nums) > 3 and len(nums[3]) == 2 else '00')
    def clamp(h, m):
        hh = max(0, min(23, int(h)))
        mm = max(0, min(59, int(m)))
        return f"{hh:02d}:{mm:02d}"
    return f"{clamp(h1,m1)} às {clamp(h2,m2)}"

def parse_data_possivel(s: str) -> str:
    if not s:
        return ''
    import re
    s = s.strip()
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r'^(\d{2})\/(\d{2})\/(\d{4})$', s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.match(r'^(\d{2})\/(\d{2})\/(\d{2})$', s)
    if m:
        return f"20{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return s

# ---------- Heurística de mapeamento ----------
KEYWORDS = {
    'disciplina': ['disciplina'],
    'descricao': ['descricao', 'descrição', 'descricao de atividades', 'descricao de atividades (descrever no minimo cinco)'],
    'nivel': ['nivel'],
    'unidade_setor': ['unidade', 'unidade/ setor', 'unidade setor', 'unidade/ setor'],
    'data_inicio': ['inicio', 'início'],
    'data_fim': ['fim'],
    'horario': ['horario', 'horário'],
    'dias_semana': ['dias da semana', 'dia', 'dias'],
    'quantidade_grupos': ['quantidade de grupos'],
    'num_estagiarios_por_grupo': ['nº de estagiarios por grupo', 'n de estagiarios por grupo', 'n de estagiarios'],
    'carga_horaria_individual': ['carga horaria individual', 'carga horaria'],
    'supervisor_nome': ['supervisor'],
    'supervisor_conselho': ['nº conselho', 'numero conselho', 'n conselho'],
    'valor': ['valor']
}

def norm(s: str | None) -> str:
    if not s:
        return ''
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFD', s)
    return ''.join(ch for ch in s if not unicodedata.combining(ch))

# ---------- Função principal de parsing ----------

def extrair_atividades(arquivo: str) -> tuple[list[dict[str, Any]], dict[str,str]]:
    atividades: list[dict[str, Any]] = []
    suggestions: dict[str, str] = {}
    rows: list[dict[str, Any]] = []

    if arquivo.lower().endswith('.csv'):
        import csv
        import io
        encodings = ['utf-8', 'latin-1', 'iso-8859-1']
        raw = None
        for enc in encodings:
            try:
                with open(arquivo, 'r', encoding=enc) as f:
                    raw = f.read()
                break
            except UnicodeDecodeError:
                continue
        if raw is None:
            raise RuntimeError('Não foi possível decodificar CSV')
        sample = raw[:1024]
        delim = ';' if sample.count(';') > sample.count(',') else ','
        reader = csv.DictReader(io.StringIO(raw), delimiter=delim)
        headers = reader.fieldnames or []
        for h in headers:
            hn = norm(h)
            for key, vars_ in KEYWORDS.items():
                if any(hn.startswith(v) for v in vars_ if hn):
                    suggestions.setdefault(key, h)
        for row in reader:
            rows.append(dict(row))
    else:
        import openpyxl, io
        from openpyxl import load_workbook
        wb = load_workbook(arquivo, data_only=True)
        sh = wb.active
        # detectar header row
        header_row_idx = 1
        detected = False
        for r in range(1, min(sh.max_row, 30) + 1):
            vals = [sh.cell(row=r, column=c).value for c in range(1, sh.max_column+1)]
            normed = [norm(v) for v in vals]
            match_count = 0
            for cell_val in normed:
                for _, variants in KEYWORDS.items():
                    if any(cell_val.startswith(v) for v in variants if cell_val):
                        match_count += 1
                        break
            if match_count >= 3:
                header_row_idx = r
                detected = True
                break
        raw_headers = [sh.cell(row=header_row_idx, column=c).value for c in range(1, sh.max_column+1)]
        headers = [str(h or f'Coluna_{i+1}') for i, h in enumerate(raw_headers)]
        for h in headers:
            hn = norm(h)
            for key, vars_ in KEYWORDS.items():
                if any(hn.startswith(v) for v in vars_ if hn):
                    suggestions.setdefault(key, h)
        # linhas
        for r in range(header_row_idx + 1, sh.max_row + 1):
            row_vals = {}
            empty = True
            for i, h in enumerate(headers):
                val = sh.cell(row=r, column=i+1).value
                sval = '' if val is None else str(val)
                if sval.strip():
                    empty = False
                row_vals[h] = sval
            if not empty:
                rows.append(row_vals)
    # Converter rows -> atividades
    for r in rows:
        a = {k: '' for k in KEYWORDS.keys()}
        for internal_key, header_name in suggestions.items():
            if header_name in r:
                a[internal_key] = r[header_name].strip() if isinstance(r[header_name], str) else r[header_name]
        # Limpeza de placeholders e espaços
        for k,v in list(a.items()):
            if isinstance(v,str):
                vv = v.strip()
                if vv.upper() == '#NOME?':
                    a[k] = ''
                else:
                    a[k] = vv
        # normalizações
        if a['horario']:
            a['horario'] = normalizar_horario(str(a['horario']))
        if a['dias_semana']:
            a['dias_semana'] = parse_dias(str(a['dias_semana']))
        if a['data_inicio']:
            a['data_inicio'] = parse_data_possivel(str(a['data_inicio']))
        if a['data_fim']:
            a['data_fim'] = parse_data_possivel(str(a['data_fim']))
        # números
        for num_field in ['quantidade_grupos','num_estagiarios_por_grupo']:
            import re
            rawv = a.get(num_field)
            if rawv is None:
                a[num_field] = None
            else:
                digits = re.sub(r'\D+', '', str(rawv))
                a[num_field] = int(digits) if digits else None
        # numero conselho: manter formato sem espaços extras (pode deixar pontos como digitado ou padronizar só dígitos)
        if a.get('supervisor_conselho'):
            import re
            digits = re.sub(r'[^0-9]', '', str(a['supervisor_conselho']))
            a['supervisor_conselho'] = digits or ''
        # filtro linha vazia
        if any((str(v).strip() if v is not None else '') for v in a.values()):
            atividades.append(a)
    return atividades, suggestions

# ---------- Operações DB ----------

def salvar_anexo2(session: Session, estagio_id: int, atividades: List[Dict[str, Any]], instituicao: str | None, curso: str | None, exercicio: str | None, status: str, logo_url: str | None, label: str | None, comment: str | None):
    existente = crud.get_anexo2_by_estagio(session, estagio_id)
    payload = schemas.Anexo2Create(
        estagio_id=estagio_id,
        instituicao_ensino=instituicao,
        curso=curso,
        exercicio=exercicio,
        status=status,
        atividades=[schemas.Anexo2AtividadeCreate(**a) for a in atividades],
        logo_url=logo_url
    )
    if existente:
        # update (mantém versao bump no crud)
        anexo = crud.update_anexo2(session, existente.id, payload)
    else:
        anexo = crud.create_anexo2(session, payload)
    # registrar versão manualmente se final e label/comment
    if status == 'final' and (label or comment):
        # replicar lógica de versionamento existente
        from datetime import datetime
        import json as _json
        versoes = crud.list_anexo2_versions(session, estagio_id)
        proxima = (versoes[0].versao + 1) if versoes else 1
        v = models.Anexo2Version(
            estagio_id=estagio_id,
            versao=proxima,
            payload=_json.dumps({
                'cabecalho': {
                    'instituicao_ensino': anexo.instituicao_ensino,
                    'curso': anexo.curso,
                    'exercicio': anexo.exercicio,
                    'status': anexo.status,
                    'logo_url': anexo.logo_url,
                },
                'atividades': [
                    {
                        'id': a.id,
                        'disciplina': a.disciplina,
                        'descricao': a.descricao,
                        'nivel': a.nivel,
                        'unidade_setor': a.unidade_setor,
                        'data_inicio': a.data_inicio,
                        'data_fim': a.data_fim,
                        'horario': a.horario,
                        'dias_semana': a.dias_semana,
                        'quantidade_grupos': a.quantidade_grupos,
                        'num_estagiarios_por_grupo': a.num_estagiarios_por_grupo,
                        'carga_horaria_individual': a.carga_horaria_individual,
                        'supervisor_nome': a.supervisor_nome,
                        'supervisor_conselho': a.supervisor_conselho,
                        'valor': a.valor,
                    } for a in anexo.atividades
                ]
            }, ensure_ascii=False),
            label=label,
            comment=comment
        )
        session.add(v)
        session.commit()
    return anexo

# ---------- CLI ----------

def main():
    ap = argparse.ArgumentParser(description='Importar planilha para Anexo II')
    ap.add_argument('--arquivo', help='Caminho do arquivo .xlsx/.xls/.csv')
    ap.add_argument('--estagio-id', type=int, help='ID do estágio destino')
    ap.add_argument('--criar-estagio', action='store_true', help='Criar estágio automaticamente se não informar --estagio-id')
    ap.add_argument('--estagio-nome', help='Nome do estágio ao criar (default: argumento --instituicao ou nome do arquivo)')
    ap.add_argument('--estagio-email', help='Email placeholder do estágio ao criar (default: estagio+timestamp@example.com)')
    ap.add_argument('--instituicao', help='Instituição de ensino (cabeçalho)')
    ap.add_argument('--curso', help='Curso (cabeçalho)')
    ap.add_argument('--exercicio', help='Exercício (ex: 2025)')
    ap.add_argument('--status', choices=['final','rascunho'], default='final')
    ap.add_argument('--logo-url', help='URL do logo (opcional)')
    ap.add_argument('--label', help='Rótulo da versão (opcional)')
    ap.add_argument('--comment', help='Comentário da versão (opcional)')
    ap.add_argument('--listar-estagios', action='store_true', help='Listar estágios existentes')
    ap.add_argument('--filtro', help='Filtro substring para nome do estágio ao listar')
    args = ap.parse_args()

    session = db.SessionLocal()
    try:
        if args.listar_estagios:
            q = session.query(models.Estagio).order_by(models.Estagio.id.asc()).all()
            for e in q:
                if args.filtro and args.filtro.lower() not in e.nome.lower():
                    continue
                print(f"{e.id:4d} | {e.nome} | curso_id={e.curso_id} | supervisor_id={e.supervisor_id}")
            return
        if not args.arquivo:
            print('Erro: informe --arquivo.')
            return
        if not args.estagio_id and not args.criar_estagio:
            print('Erro: informe --estagio-id ou use --criar-estagio para criação automática.')
            return
        # Criar estágio se solicitado
        estagio_id = args.estagio_id
        if not estagio_id and args.criar_estagio:
            from datetime import datetime
            nome_estagio = args.estagio_nome or args.instituicao or os.path.splitext(os.path.basename(args.arquivo))[0]
            email_estagio = args.estagio_email or f"estagio+{int(datetime.utcnow().timestamp())}@example.com"
            # Supervisor: pegar primeiro existente ou criar dummy
            sup = session.query(models.Supervisor).first()
            if not sup:
                from app import schemas as _schemas, crud as _crud  # tipo ignore
                dummy = _schemas.SupervisorCreate(nome='Supervisor Padrão', email=f'supervisor+{int(datetime.utcnow().timestamp())}@example.com')
                sup = _crud.create_supervisor(session, dummy)  # type: ignore
            novo = models.Estagio(
                nome=nome_estagio,
                email=email_estagio,
                telefone='',
                periodo='',
                supervisor_id=sup.id,
            )
            session.add(novo)
            session.commit()
            session.refresh(novo)
            estagio_id = novo.id
            print(f"🆕 Estágio criado: id={estagio_id} nome='{nome_estagio}' supervisor_id={sup.id}")
        if not os.path.isfile(args.arquivo):
            print(f"Arquivo não encontrado: {args.arquivo}")
            return
        print(f"📥 Lendo arquivo: {args.arquivo}")
        atividades, sugest = extrair_atividades(args.arquivo)
        if not atividades:
            print('Nenhuma atividade detectada.')
            return
        print(f"✔ Detectadas {len(atividades)} atividades. Sugestões de mapeamento: {json.dumps(sugest, ensure_ascii=False)}")
        anexo = salvar_anexo2(
            session,
            estagio_id=estagio_id,
            atividades=atividades,
            instituicao=args.instituicao,
            curso=args.curso,
            exercicio=args.exercicio,
            status=args.status,
            logo_url=args.logo_url,
            label=args.label,
            comment=args.comment
        )
        print(f"✅ Anexo II salvo (estagio_id={anexo.estagio_id}). Total atividades armazenadas: {len(anexo.atividades)}")
    finally:
        session.close()

if __name__ == '__main__':
    main()

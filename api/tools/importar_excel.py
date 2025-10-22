#!/usr/bin/env python3
"""
Script para importar dados do Excel da Faculdade Guarapuava para o sistema de estágios
"""
import openpyxl
import sys
import os
from pathlib import Path
from datetime import datetime, date, time
from decimal import Decimal
import re

# Adicionar o diretório pai ao path para importar módulos do sistema
sys.path.append(str(Path(__file__).parent.parent))

from app.db import SessionLocal
from app.models import Estagio, Curso, Supervisor, Unidade, Territorio, Instituicao

def parse_data(data_str):
    """Converte string de data para objeto date"""
    if not data_str or data_str.strip() == "":
        return None
    
    # Tentar diferentes formatos de data
    formatos = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']
    
    for formato in formatos:
        try:
            return datetime.strptime(data_str.strip(), formato).date()
        except ValueError:
            continue
    
    print(f"⚠️ Não foi possível converter data: {data_str}")
    return None

def parse_horario(horario_str):
    """Converte string de horário para objeto time"""
    if not horario_str or horario_str.strip() == "":
        return None
    
    # Tentar diferentes formatos de horário
    formatos = ['%H:%M', '%H:%M:%S']
    
    for formato in formatos:
        try:
            return datetime.strptime(horario_str.strip(), formato).time()
        except ValueError:
            continue
    
    print(f"⚠️ Não foi possível converter horário: {horario_str}")
    return None

def parse_valor(valor_str):
    """Converte string de valor para float"""
    if not valor_str or valor_str.strip() == "":
        return None
    
    # Remover símbolos de moeda e espaços
    valor_limpo = re.sub(r'[R$\s]', '', str(valor_str))
    valor_limpo = valor_limpo.replace(',', '.')
    
    try:
        return float(valor_limpo)
    except ValueError:
        print(f"⚠️ Não foi possível converter valor: {valor_str}")
        return None

def importar_excel(arquivo_excel, instituicao_nome="Faculdade Guarapuava"):
    """Importa dados do Excel para o banco de dados"""
    
    if not os.path.exists(arquivo_excel):
        print(f"❌ Arquivo não encontrado: {arquivo_excel}")
        return False
    
    try:
        wb = openpyxl.load_workbook(arquivo_excel)
        db = SessionLocal()
        
        # Buscar ou criar instituição
        instituicao = db.query(Instituicao).filter(Instituicao.nome == instituicao_nome).first()
        if not instituicao:
            instituicao = Instituicao(nome=instituicao_nome)
            db.add(instituicao)
            db.commit()
            print(f"✅ Instituição criada: {instituicao_nome}")
        
        total_importados = 0
        
        for sheet_name in wb.sheetnames:
            print(f"\n📋 Processando planilha: {sheet_name}")
            ws = wb[sheet_name]
            
            # Buscar ou criar curso
            curso = db.query(Curso).filter(Curso.nome == sheet_name).first()
            if not curso:
                curso = Curso(nome=sheet_name)
                db.add(curso)
                db.commit()
                print(f"✅ Curso criado: {sheet_name}")
            
            # Linha 6 contém os cabeçalhos
            header_row = 6
            
            # Processar dados a partir da linha 8 (após cabeçalhos e linha de início/fim)
            linhas_processadas = 0
            
            for row in range(8, ws.max_row + 1):
                # Extrair dados da linha
                disciplina = ws.cell(row, 1).value
                descricao = ws.cell(row, 2).value
                nivel = ws.cell(row, 3).value
                unidade_nome = ws.cell(row, 4).value
                data_inicio_str = ws.cell(row, 5).value
                data_fim_str = ws.cell(row, 6).value
                horario_str = ws.cell(row, 7).value
                dias_semana = ws.cell(row, 8).value
                quantidade_grupos = ws.cell(row, 9).value
                num_estagiarios = ws.cell(row, 10).value
                carga_horaria = ws.cell(row, 11).value
                supervisor_nome = ws.cell(row, 12).value
                numero_conselho = ws.cell(row, 13).value
                valor = ws.cell(row, 14).value
                
                # Verificar se há dados suficientes na linha
                if not any([disciplina, unidade_nome, supervisor_nome]):
                    continue  # Pular linhas vazias
                
                # Buscar ou criar supervisor
                supervisor = None
                if supervisor_nome:
                    supervisor = db.query(Supervisor).filter(Supervisor.nome == str(supervisor_nome).strip()).first()
                    if not supervisor:
                        # Criar supervisor básico (email será necessário ajustar depois)
                        email_supervisor = f"{str(supervisor_nome).lower().replace(' ', '.')}@estagios.local"
                        supervisor = Supervisor(
                            nome=str(supervisor_nome).strip(),
                            email=email_supervisor,
                            numero_conselho=str(numero_conselho).strip() if numero_conselho else None
                        )
                        db.add(supervisor)
                        db.commit()
                        print(f"✅ Supervisor criado: {supervisor_nome}")
                
                # Buscar ou criar unidade
                unidade = None
                if unidade_nome:
                    unidade = db.query(Unidade).filter(Unidade.nome == str(unidade_nome).strip()).first()
                    if not unidade:
                        unidade = Unidade(nome=str(unidade_nome).strip())
                        db.add(unidade)
                        db.commit()
                        print(f"✅ Unidade criada: {unidade_nome}")
                
                # Criar estágio
                estagio = Estagio(
                    nome=f"Estágio {disciplina or 'Sem nome'} - {sheet_name}",
                    email="nao.informado@estagios.local",  # Email padrão
                    disciplina=str(disciplina).strip() if disciplina else None,
                    nivel=str(nivel).strip() if nivel else None,
                    data_inicio=parse_data(str(data_inicio_str)) if data_inicio_str else None,
                    data_fim=parse_data(str(data_fim_str)) if data_fim_str else None,
                    horario_inicio=parse_horario(str(horario_str).split('-')[0]) if horario_str and '-' in str(horario_str) else None,
                    horario_fim=parse_horario(str(horario_str).split('-')[1]) if horario_str and '-' in str(horario_str) else None,
                    dias_semana=str(dias_semana).strip() if dias_semana else None,
                    quantidade_grupos=int(quantidade_grupos) if quantidade_grupos and str(quantidade_grupos).isdigit() else 1,
                    num_estagiarios=int(num_estagiarios) if num_estagiarios and str(num_estagiarios).isdigit() else None,
                    carga_horaria=int(carga_horaria) if carga_horaria and str(carga_horaria).isdigit() else None,
                    valor_total=parse_valor(valor) if valor else None,
                    observacoes=str(descricao).strip() if descricao else None,
                    supervisor_id=supervisor.id if supervisor else None,
                    unidade_id=unidade.id if unidade else None,
                    curso_id=curso.id,
                    instituicao_id=instituicao.id,
                    status="ativo"
                )
                
                db.add(estagio)
                linhas_processadas += 1
            
            if linhas_processadas > 0:
                db.commit()
                print(f"✅ {linhas_processadas} estágios importados da planilha {sheet_name}")
                total_importados += linhas_processadas
            else:
                print(f"⚠️ Nenhum dado encontrado na planilha {sheet_name}")
        
        db.close()
        print(f"\n🎉 Importação concluída! Total de estágios importados: {total_importados}")
        return True
        
    except Exception as e:
        print(f"❌ Erro durante importação: {e}")
        if 'db' in locals():
            db.rollback()
            db.close()
        return False

if __name__ == "__main__":
    arquivo = "FACULDADE GUARAPUAVA (2).xlsx"
    
    if len(sys.argv) > 1:
        arquivo = sys.argv[1]
    
    print(f"🚀 Iniciando importação do arquivo: {arquivo}")
    sucesso = importar_excel(arquivo)
    
    if sucesso:
        print("✅ Importação realizada com sucesso!")
    else:
        print("❌ Falha na importação!")
        sys.exit(1)
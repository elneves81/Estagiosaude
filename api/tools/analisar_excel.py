#!/usr/bin/env python3
"""
Analisar o arquivo Excel da Faculdade Guarapuava
"""
import openpyxl
import json
from pathlib import Path

def analisar_excel():
    """Analisa estrutura e conteúdo do arquivo Excel"""
    arquivo = Path(__file__).parent / "FACULDADE GUARAPUAVA (2).xlsx"
    
    try:
        wb = openpyxl.load_workbook(arquivo)
        print(f"📋 Planilhas encontradas: {wb.sheetnames}")
        
        resultados = {}
        
        for sheet_name in wb.sheetnames:
            print(f"\n🔍 Analisando planilha: {sheet_name}")
            ws = wb[sheet_name]
            
            if ws.max_row < 2:
                print("  ⚠️ Planilha vazia ou só com cabeçalho")
                continue
                
            # Encontrar linha de cabeçalhos (procurar por palavras-chave)
            header_row = None
            for row in range(1, min(10, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    val = ws.cell(row, col).value
                    if val and isinstance(val, str):
                        val_lower = val.lower()
                        if any(keyword in val_lower for keyword in ['nome', 'curso', 'data', 'local', 'supervisor']):
                            header_row = row
                            break
                if header_row:
                    break
            
            if not header_row:
                print("  ⚠️ Não encontrei linha de cabeçalhos clara")
                continue
                
            # Extrair cabeçalhos
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(header_row, col).value
                if val:
                    headers.append(str(val).strip())
                else:
                    headers.append(f"Col_{col}")
                    
            print(f"  📝 Cabeçalhos (linha {header_row}): {headers[:10]}")
            
            # Extrair algumas linhas de dados
            dados_amostra = []
            for row in range(header_row + 1, min(header_row + 4, ws.max_row + 1)):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    val = ws.cell(row, col).value
                    if val is not None:
                        row_data[header] = str(val).strip()
                if row_data:  # só adicionar se tiver dados
                    dados_amostra.append(row_data)
            
            print(f"  📊 Total de linhas: {ws.max_row}")
            print(f"  📋 Amostra de dados:")
            for i, row_data in enumerate(dados_amostra, 1):
                print(f"    Linha {i}: {dict(list(row_data.items())[:5])}")
                
            resultados[sheet_name] = {
                'headers': headers,
                'total_rows': ws.max_row,
                'header_row': header_row,
                'sample_data': dados_amostra
            }
        
        # Análise de impacto no sistema
        print("\n🎯 ANÁLISE DE IMPACTO NO SISTEMA:")
        for sheet_name, info in resultados.items():
            print(f"\n📋 {sheet_name}:")
            headers = [h.lower() for h in info['headers']]
            
            # Mapear para campos do sistema
            mapeamento = {}
            if any('nome' in h for h in headers):
                mapeamento['nome_estagiario'] = [h for h in info['headers'] if 'nome' in h.lower()]
            if any('curso' in h for h in headers):
                mapeamento['curso'] = [h for h in info['headers'] if 'curso' in h.lower()]
            if any('local' in h or 'unidade' in h for h in headers):
                mapeamento['local/unidade'] = [h for h in info['headers'] if 'local' in h.lower() or 'unidade' in h.lower()]
            if any('supervisor' in h or 'preceptor' in h for h in headers):
                mapeamento['supervisor'] = [h for h in info['headers'] if 'supervisor' in h.lower() or 'preceptor' in h.lower()]
            if any('data' in h or 'periodo' in h for h in headers):
                mapeamento['datas'] = [h for h in info['headers'] if 'data' in h.lower() or 'periodo' in h.lower()]
                
            if mapeamento:
                print(f"  ✅ Campos mapeáveis: {mapeamento}")
            else:
                print(f"  ⚠️ Não identifiquei campos padrão de estágio")
                
        return resultados
        
    except Exception as e:
        print(f"❌ Erro: {e}")
        return None

if __name__ == "__main__":
    analisar_excel()